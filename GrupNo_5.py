"""
============================================================
  Authors: 082040014 Meltem Nur Yılmaz
           082140019 Öznur Çağdaş
           082240067 Ahmet İşbilen
  Date: 05.06.2026
  Version: 3.14
  
  Proje: Tasarruf Planı ve Borç Azaltma Modeli
  
  Gerçek Takvim, Act/365, Kesin Hassasiyet (Decimal)
  GÜNCEL: Dinamik Çözünürlük/Hizalama, Otomatik Tam Ekran,
  Bakiye Tüketim Mantığı, Üst Seviye Uyarı Sistemi,
  Dinamik Metin Kaydırma, Düzenlenebilir Çekim Tarihi,
  Akıllı Satır Gruplama, Kullanıcı İşlemi Görünürlüğü,
  Borç Yeniden Yapılandırma (Ara Ödemeye Bağlı),
  Akıllı Sekme Panel Kontrolü.
============================================================
"""

import tkinter as tk
from tkinter import ttk, filedialog
import math
import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
import calendar
import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

getcontext().prec = 28


def stopaj_hesapla(vade_gun):
    """Vade günü sayısına göre yasal stopaj oranını döndürür."""
    if vade_gun < 182:
        return Decimal('17.5')
    elif vade_gun <= 365:
        return Decimal('15.0')
    elif vade_gun <= 730:
        return Decimal('12.0')
    else:
        return Decimal('10.0')


# ─────────────────────────── Renk Paletleri ───────────────────────────
TEMA_KARANLIK = {
    "DARK_BG": "#0D1117", "PANEL_BG": "#161B22", "BORDER": "#30363D",
    "ACCENT_BLUE": "#58A6FF", "ACCENT_DARK_BLUE": "#1F6FEB", "ACCENT_GREEN": "#3FB950", "ACCENT_RED": "#F85149",
    "ACCENT_GOLD": "#D29922", "ACCENT_ORANGE": "#E8820C", "TEXT_PRIMARY": "#E6EDF3",
    "TEXT_MUTED": "#8B949E", "CHART_BG": "#0D1117", "ENTRY_BG": "#1C2128",
    "TEXTBOX_BG": "#88A7B4", "TEXTBOX_FG": "#000000",
    "ACCENT_TAHAKKUK": "#58A6FF", "ROW_EVEN": "#161B22", "ROW_ODD": "#0D1117"
}

TEMA_AYDINLIK = {
    "DARK_BG": "#F0F4F8", "PANEL_BG": "#FFFFFF", "BORDER": "#CBD5E1",
    "ACCENT_BLUE": "#1A73E8", "ACCENT_DARK_BLUE": "#0D47A1", "ACCENT_GREEN": "#035F46", "ACCENT_RED": "#B91C1C",
    "ACCENT_GOLD": "#B45309", "ACCENT_ORANGE": "#EA580C", "TEXT_PRIMARY": "#0F172A",
    "TEXT_MUTED": "#475569", "CHART_BG": "#F0F4F8", "ENTRY_BG": "#F1F5F9",
    "TEXTBOX_BG": "#FFFFFF", "TEXTBOX_FG": "#0F172A",
    "ACCENT_TAHAKKUK": "#083D7F", "ROW_EVEN": "#FFFFFF", "ROW_ODD": "#a4c7d5"
}


def para_format(deger):
    return Decimal(str(deger)).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)


# ══════════════════════════════════════════════════════════════════════
#  CUSTOM ENTRY FABRİKA FONKSİYONU
# ══════════════════════════════════════════════════════════════════════
def CustomEntry(parent, is_dark, **kwargs):
    disabled_bg = kwargs.pop('disabledbackground', '#b0b0b0')
    disabled_fg = kwargs.pop('disabledforeground', '#3A3A3A')
    readonly_bg = kwargs.pop('readonlybackground', None)

    kwargs.pop('highlightthickness', None)
    kwargs.pop('highlightbackground', None)
    kwargs.pop('highlightcolor', None)

    if is_dark:
        outer_frame = tk.Frame(parent, bg="#000000", padx=1, pady=1)
        entry_bg = kwargs.get('bg', '#1C2128')
        inner_frame = tk.Frame(outer_frame, bg=entry_bg, padx=1, pady=1)
        inner_frame.pack(fill="both", expand=True)

        kwargs['relief'] = 'flat'
        kwargs['bd'] = 0

        entry = tk.Entry(inner_frame, **kwargs)
        entry.pack(fill="both", expand=True)

        entry.pack_orig = entry.pack
        entry.grid_orig = entry.grid
        entry.pack = lambda **kw: outer_frame.pack(**kw)
        entry.grid = lambda **kw: outer_frame.grid(**kw)
        entry.pack_forget = lambda: outer_frame.pack_forget()
        entry.grid_remove = lambda: outer_frame.grid_remove()
    else:
        kwargs['relief'] = 'solid'
        kwargs['bd'] = 1
        entry = tk.Entry(parent, **kwargs)

    entry.config(disabledbackground=disabled_bg, disabledforeground=disabled_fg)
    if readonly_bg:
        entry.config(readonlybackground=readonly_bg)
    return entry


# ══════════════════════════════════════════════════════════════════════
#  TOOLTIP SINIFI
# ══════════════════════════════════════════════════════════════════════
class Tooltip:
    def __init__(self, widget, text, r):
        self.widget = widget
        self.text = text
        self.r = r
        self.tip_window = None
        self.id = None
        self.widget.bind("<Enter>", self.enter, add="+")
        self.widget.bind("<Leave>", self.leave, add="+")

    def enter(self, event=None):
        self.schedule()

    def leave(self, event=None):
        self.unschedule()
        self.hide_tip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(300, self.show_tip)

    def unschedule(self):
        id_ = self.id
        self.id = None
        if id_: self.widget.after_cancel(id_)

    def show_tip(self, event=None):
        if self.tip_window or not self.text: return
        x, y, _, _ = self.widget.bbox("insert") if hasattr(self.widget, "bbox") else (0, 0, 0, 0)
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        underline_words = getattr(self, "underline_words", [])
        if underline_words:
            font_normal = ("Courier New", 10)
            font_underline = ("Courier New", 10, "underline")
            txt = tk.Text(tw, background=self.r["ENTRY_BG"], foreground=self.r["TEXT_PRIMARY"], relief="solid",
                          borderwidth=1, font=font_normal, wrap="none", state="normal", cursor="arrow", padx=6, pady=4)
            txt.tag_configure("underline", font=font_underline)
            remaining = self.text
            while remaining:
                earliest_idx = len(remaining)
                earliest_word = None
                for w in underline_words:
                    idx = remaining.find(w)
                    if idx != -1 and idx < earliest_idx: earliest_idx = idx; earliest_word = w
                if earliest_word is None: txt.insert("end", remaining); break
                txt.insert("end", remaining[:earliest_idx])
                txt.insert("end", earliest_word, "underline")
                remaining = remaining[earliest_idx + len(earliest_word):]
            txt.config(state="disabled")
            lines = self.text.split("\n")
            max_chars = max(len(l) for l in lines)
            txt.config(width=max_chars, height=len(lines))
            txt.pack()
        else:
            label = tk.Label(tw, text=self.text, justify="left", background=self.r["ENTRY_BG"],
                             foreground=self.r["TEXT_PRIMARY"], relief="solid", borderwidth=1, font=("Courier New", 10))
            label.pack(ipadx=6, ipady=4)

    def hide_tip(self):
        tw = self.tip_window
        self.tip_window = None
        if tw: tw.destroy()


# ══════════════════════════════════════════════════════════════════════
#  MATEMATİK KATMANI
# ══════════════════════════════════════════════════════════════════════
def hesapla_mevduat(baslangic, yillik_faiz, aylik_yatirim, donem_sayisi, stopaj_orani, vade_gun, baslangic_tarihi,
                    erken_cekim_var=False, tek_seferlik=False, ozel_islemler=None, enflasyon=0.0,
                    cekim_tarihleri=None, erken_cekim_donemler=None):
    if ozel_islemler is None: ozel_islemler = {}
    if cekim_tarihleri is None: cekim_tarihleri = {}
    if erken_cekim_donemler is None: erken_cekim_donemler = set()
    A0 = para_format(baslangic)
    F = Decimal(str(yillik_faiz)) / Decimal('100')
    oto_stopaj = stopaj_hesapla(vade_gun)
    SC = Decimal('1') - (oto_stopaj / Decimal('100'))
    duzenli_yatirim = para_format(aylik_yatirim)
    gun = Decimal(str(vade_gun))
    enf_orani = float(enflasyon) / 100.0
    r_donem = F * gun / Decimal('365')
    p_min = para_format(A0 * r_donem * SC)

    bakiyeler_nominal = [float(A0)]
    bakiyeler_reel = [float(A0)]
    ekstre = []
    bakiye = A0
    aktif_tarih = baslangic_tarihi

    vade_bozuldu = False
    vade_bozulan_donemler = set()
    bakiye_asimi_var = False
    bakiye_asimi_detay = []

    yanan_faiz_toplami = Decimal('0.00')
    toplam_brut_faiz = Decimal('0.00')
    toplam_stopaj_kesinti = Decimal('0.00')
    toplam_net_faiz = Decimal('0.00')

    _sonraki_baslangic = None

    for n in range(donem_sayisi):
        donem_no = n + 1
        if _sonraki_baslangic is not None:
            aktif_tarih = _sonraki_baslangic + datetime.timedelta(days=int(vade_gun))
            _sonraki_baslangic = None
        else:
            aktif_tarih += datetime.timedelta(days=int(vade_gun))

        brut_faiz = para_format(bakiye * r_donem)
        net_faiz = para_format(brut_faiz * SC)
        stopaj_kesinti = para_format(brut_faiz - net_faiz)
        ozel_tutar = ozel_islemler.get(donem_no, Decimal('0.00'))
        uygulanacak_islem = Decimal('0.00')

        if not tek_seferlik or (tek_seferlik and donem_no == 1): uygulanacak_islem = duzenli_yatirim

        donem_vade_bozuldu = False

        if erken_cekim_var and donem_no in erken_cekim_donemler:
            vade_bozuldu = True
            donem_vade_bozuldu = True
            vade_bozulan_donemler.add(donem_no)
            yanan_faiz_toplami += net_faiz
            net_faiz = Decimal('0.00')
            stopaj_kesinti = Decimal('0.00')
            brut_faiz = Decimal('0.00')
            if donem_no in cekim_tarihleri:
                _sonraki_baslangic = cekim_tarihleri[donem_no]

        toplam_brut_faiz += brut_faiz
        toplam_stopaj_kesinti += stopaj_kesinti
        toplam_net_faiz += net_faiz
        bakiye += net_faiz

        nakit_girisi = (uygulanacak_islem if uygulanacak_islem > Decimal('0') else Decimal('0.00')) + \
                       (ozel_tutar if ozel_tutar > Decimal('0') else Decimal('0.00'))

        istenen_cikis = abs(uygulanacak_islem if uygulanacak_islem < Decimal('0') else Decimal('0.00')) + \
                        abs(ozel_tutar if ozel_tutar < Decimal('0') else Decimal('0.00'))

        bakiye += nakit_girisi

        gerceklesen_cikis = istenen_cikis
        if istenen_cikis > bakiye:
            gerceklesen_cikis = bakiye
            bakiye_asimi_var = True
            bakiye_asimi_detay.append({
                "donem": donem_no,
                "istenen": istenen_cikis,
                "cekilen": gerceklesen_cikis
            })

        bakiye -= gerceklesen_cikis
        guncel_islem = nakit_girisi - gerceklesen_cikis

        if bakiye < Decimal('0'): bakiye = Decimal('0')

        bakiyeler_nominal.append(float(para_format(bakiye)))
        gecen_gun = donem_no * float(vade_gun)
        reel_bakiye = bakiye * Decimal(str(math.pow(1.0 + enf_orani, -gecen_gun / 365.0))) if enf_orani > 0 else bakiye
        bakiyeler_reel.append(float(para_format(reel_bakiye)))

        kullanici_islemi_var = (donem_no in ozel_islemler) or donem_vade_bozuldu

        ekstre.append(
            {"donem": donem_no,
             "baslangic_tarih": (aktif_tarih - datetime.timedelta(days=int(vade_gun))).strftime("%d.%m.%Y"),
             "tarih": aktif_tarih.strftime("%d.%m.%Y"), "faiz": net_faiz, "islem": guncel_islem,
             "bakiye": bakiye, "reel_bakiye": para_format(reel_bakiye), "vade_bozuldu": donem_vade_bozuldu,
             "kullanici_islemi": kullanici_islemi_var})

    if bakiye <= Decimal('0'):
        karakteristik = "YAKINSAK\n(Sıfıra Eriyip Bitti)"
    elif tek_seferlik:
        karakteristik = "IRAKSAK\n(Tek Seferlik Yatırım, Faiz Birikir)"
    elif duzenli_yatirim >= Decimal('0'):
        karakteristik = "IRAKSAK\n(Sürekli Büyür)"
    elif abs(duzenli_yatirim) > para_format(A0 * r_donem * SC):
        karakteristik = "YAKINSAK\n(Zamanla Eriyip Biter)"
    else:
        karakteristik = "IRAKSAK\n(Faiz Çekimi Karşılıyor, Para Bitmez)"

    return {"nominal": bakiyeler_nominal, "reel": bakiyeler_reel, "p_min": p_min, "bitis_tarihi": aktif_tarih,
            "vade_bozuldu": vade_bozuldu, "vade_bozulan_donemler": vade_bozulan_donemler,
            "bakiye_asimi_var": bakiye_asimi_var, "bakiye_asimi_detay": bakiye_asimi_detay,
            "yanan_faiz": yanan_faiz_toplami, "ekstre": ekstre, "karakteristik": karakteristik,
            "toplam_brut_faiz": para_format(toplam_brut_faiz),
            "toplam_stopaj_kesinti": para_format(toplam_stopaj_kesinti),
            "toplam_net_faiz": para_format(toplam_net_faiz), "oto_stopaj": oto_stopaj}


def hesapla_borc(baslangic, yillik_faiz, aylik_odeme, vergi_orani, baslangic_tarihi, ara_odemeler=None,
                 hedef_vade=None, yeniden_yapilandir=False):
    if ara_odemeler is None: ara_odemeler = {}
    B0 = para_format(baslangic)
    F = Decimal(str(yillik_faiz)) / Decimal('100')
    vergi_orani_d = Decimal(str(vergi_orani)) / Decimal('100')
    r_aylik = (F / Decimal('12')) * (Decimal('1') + vergi_orani_d)
    p_min = para_format(B0 * r_aylik)

    P = para_format(aylik_odeme)
    bakiyeler_nom = [float(B0)]
    ekstre = []
    borc = B0
    aktif_tarih = baslangic_tarihi
    toplam_odenen_nom = Decimal('0.00')
    iraksar = False
    dongu_limiti = hedef_vade if hedef_vade is not None else 600

    toplam_ara = sum(ara_odemeler.values()) if ara_odemeler else Decimal("0.00")
    ara_yeterli = toplam_ara >= B0

    if P <= p_min and not ara_yeterli and hedef_vade is None:
        iraksar = True
        dongu_limiti = 24
        if P < p_min:
            iraksar_tip = "buyuyor"
        else:
            iraksar_tip = "sabit"
    else:
        iraksar_tip = "kapaniyor"

    # Hedef vade verilmediyse standart bitiş ayını doğal vade olarak hesapla
    hedef_vade_calc = hedef_vade
    if yeniden_yapilandir and hedef_vade is None:
        if P > p_min:
            if r_aylik > 0:
                try:
                    hedef_vade_calc = math.ceil(-math.log(1 - float(B0 * r_aylik / P)) / math.log(1 + float(r_aylik)))
                except:
                    hedef_vade_calc = 600
            else:
                hedef_vade_calc = math.ceil(float(B0 / P))
        else:
            hedef_vade_calc = 600

    n = 0
    capa_gun = baslangic_tarihi.day
    vade_kurtarmiyor = False
    kalan_borc_gosterim = Decimal('0.00')

    while n < dongu_limiti and borc > Decimal('0'):
        eski_tarih = aktif_tarih
        yeni_ay = eski_tarih.month % 12 + 1
        yeni_yil = eski_tarih.year + (eski_tarih.month // 12)
        try:
            aktif_tarih = datetime.date(yeni_yil, yeni_ay, capa_gun)
        except ValueError:
            aktif_tarih = datetime.date(yeni_yil, yeni_ay, calendar.monthrange(yeni_yil, yeni_ay)[1])

        donem_faizi = para_format(borc * r_aylik)
        ekstra = ara_odemeler.get(n + 1, Decimal('0.00'))
        bu_ay_odenen_plan = P + ekstra
        faizli_borc = borc + donem_faizi

        kalan_fark = faizli_borc - bu_ay_odenen_plan

        if hedef_vade is not None and (n + 1) == hedef_vade and kalan_fark > Decimal('1.00'):
            vade_kurtarmiyor = True
            kalan_borc_gosterim = kalan_fark

        son_taksit_mi = (
                (hedef_vade is not None and (n + 1) == hedef_vade)
                or faizli_borc <= bu_ay_odenen_plan
                or (Decimal('0') < kalan_fark < Decimal('1.00'))
        )

        if son_taksit_mi:
            odenen_anapara = borc
            toplam_odenen_nom += faizli_borc
            gercek_odeme = faizli_borc
            borc = Decimal('0.00')
        else:
            odenen_anapara = bu_ay_odenen_plan - donem_faizi
            toplam_odenen_nom += bu_ay_odenen_plan
            gercek_odeme = bu_ay_odenen_plan
            borc = faizli_borc - bu_ay_odenen_plan

        bakiyeler_nom.append(float(borc))
        ekstre.append(
            {"taksit": n + 1, "tarih": aktif_tarih.strftime("%d.%m.%Y"), "kalan_borc": borc, "faiz": donem_faizi,
             "anapara": odenen_anapara, "ara_odeme": ekstra, "tutar": gercek_odeme})

        # --- YENİ EKLENEN KISIM: Taksit Düşürme (Yeniden Yapılandırma) ---
        if yeniden_yapilandir and ekstra > Decimal('0') and borc > Decimal('0'):
            if hedef_vade_calc is not None and (n + 1) < hedef_vade_calc:
                kalan_ay = hedef_vade_calc - (n + 1)
                if r_aylik > 0:
                    factor = Decimal('1') - (Decimal('1') + r_aylik) ** (-kalan_ay)
                    P = (borc * r_aylik / factor).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)
                else:
                    P = (borc / Decimal(str(kalan_ay))).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)

        n += 1

    karakteristik = "IRAKSAK\n(Borç Asla Kapanmaz, Büyür)" if iraksar else "YAKINSAK\n(Limit Sınırında Borç Sıfırlanır)"
    return {"nominal": bakiyeler_nom, "iraksar": iraksar, "iraksar_tip": iraksar_tip if iraksar else "kapaniyor",
            "kapanma_ay": n, "p_min": p_min,
            "toplam_odenen_nom": toplam_odenen_nom, "ekstre": ekstre, "bitis_tarihi": aktif_tarih,
            "karakteristik": karakteristik, "vade_kurtarmiyor": vade_kurtarmiyor,
            "kalan_borc": kalan_borc_gosterim if vade_kurtarmiyor else borc}


# ══════════════════════════════════════════════════════════════════════
#  TAKVİM POPUP & GRAFİK KATMANI
# ══════════════════════════════════════════════════════════════════════
class TakvimPopup(tk.Toplevel):
    def __init__(self, parent, tarih_var, renkler, entry_widget, min_tarih=None):
        super().__init__(parent)
        self.tarih_var = tarih_var
        self.renkler = renkler
        self.min_tarih = min_tarih
        self.overrideredirect(True)
        self.configure(bg=renkler["BORDER"])
        try:
            mevcut = datetime.datetime.strptime(tarih_var.get(), "%d.%m.%Y").date()
        except:
            mevcut = datetime.date.today()
        self.goruntulenen_yil, self.goruntulenen_ay, self.secili_tarih = mevcut.year, mevcut.month, mevcut
        self._aralik_guncelle()
        self.update_idletasks()
        self.geometry(f"+{entry_widget.winfo_rootx()}+{entry_widget.winfo_rooty() + entry_widget.winfo_height() + 2}")
        self.bind("<FocusOut>", lambda e: self.after(100, self._guvenli_kapat))
        self.focus_set()

    def _guvenli_kapat(self):
        try:
            if self.winfo_exists(): self.destroy()
        except:
            pass

    def _aralik_guncelle(self):
        for w in self.winfo_children(): w.destroy()
        r = self.renkler
        f = tk.Frame(self, bg=r["PANEL_BG"], padx=4, pady=4)
        f.pack(fill="both", expand=True, padx=1, pady=1)
        baslik = tk.Frame(f, bg=r["PANEL_BG"])
        baslik.pack(fill="x", pady=(0, 4))
        tk.Button(baslik, text="◀", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], relief="flat",
                  command=self._onceki_ay).pack(side="left")
        ay_isim = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran", "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım",
                   "Aralık"]
        tk.Label(baslik, text=f"{ay_isim[self.goruntulenen_ay - 1]} {self.goruntulenen_yil}", bg=r["PANEL_BG"],
                 fg=r["TEXT_PRIMARY"], font=("Courier New", 10, "bold")).pack(side="left", expand=True)
        tk.Button(baslik, text="▶", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], relief="flat",
                  command=self._sonraki_ay).pack(side="right")
        takvim_frame = tk.Frame(f, bg=r["PANEL_BG"])
        takvim_frame.pack()
        for i, g in enumerate(["Pt", "Sa", "Ça", "Pe", "Cu", "Ct", "Pa"]): tk.Label(takvim_frame, text=g,
                                                                                    bg=r["PANEL_BG"],
                                                                                    fg=r["ACCENT_RED"] if i == 6 else r[
                                                                                        "TEXT_MUTED"],
                                                                                    font=("Courier New", 8, "bold"),
                                                                                    width=3).grid(row=0, column=i)
        ilk_gun, ay_gunu = calendar.monthrange(self.goruntulenen_yil, self.goruntulenen_ay)
        bugun, satir, sutun = datetime.date.today(), 1, ilk_gun
        for gun in range(1, ay_gunu + 1):
            tarih = datetime.date(self.goruntulenen_yil, self.goruntulenen_ay, gun)
            devre_disi = self.min_tarih is not None and tarih <= self.min_tarih
            bg = r["ACCENT_BLUE"] if tarih == self.secili_tarih else (
                r["ACCENT_GOLD"] if tarih == bugun else r["PANEL_BG"])
            if devre_disi:
                fg = r["BORDER"]
                bg = r["PANEL_BG"]
            else:
                fg = r["DARK_BG"] if tarih in (self.secili_tarih, bugun) else (
                    r["ACCENT_RED"] if sutun == 6 else r["TEXT_PRIMARY"])
            btn = tk.Button(takvim_frame, text=str(gun), width=3, bg=bg, fg=fg, relief="flat", font=("Courier New", 9),
                            command=(lambda t=tarih: self._sec(t)) if not devre_disi else lambda: None,
                            cursor="arrow" if devre_disi else "hand2",
                            state="normal")
            btn.grid(row=satir, column=sutun, padx=1, pady=1)
            sutun += 1
            if sutun == 7: sutun, satir = 0, satir + 1

    def _onceki_ay(self):
        if self.goruntulenen_ay == 1:
            self.goruntulenen_ay, self.goruntulenen_yil = 12, self.goruntulenen_yil - 1
        else:
            self.goruntulenen_ay -= 1
        self._aralik_guncelle()

    def _sonraki_ay(self):
        if self.goruntulenen_ay == 12:
            self.goruntulenen_ay, self.goruntulenen_yil = 1, self.goruntulenen_yil + 1
        else:
            self.goruntulenen_ay += 1
        self._aralik_guncelle()

    def _sec(self, tarih):
        self.tarih_var.set(tarih.strftime("%d.%m.%Y"));
        self.destroy()


def ciz_grafik(canvas: tk.Canvas, veriler_nom: list, veriler_reel: list, renk_nom: str, baslik: str, r: dict,
               birim: str = "₺", tree=None, vade_bozulan_donemler=None, log_olcek=False):
    if vade_bozulan_donemler is None: vade_bozulan_donemler = set()
    canvas.delete("all")
    W = canvas.winfo_width() if canvas.winfo_width() > 50 else 750
    H = canvas.winfo_height() if canvas.winfo_height() > 50 else 300
    pad_left, pad_right, pad_top, pad_bottom = 85, 30, 50, 60
    iw, ih = W - pad_left - pad_right, H - pad_top - pad_bottom
    canvas.configure(bg=r["CHART_BG"])
    canvas.create_rectangle(pad_left, pad_top, pad_left + iw, pad_top + ih, fill=r["PANEL_BG"], outline=r["BORDER"])

    tum_veriler = veriler_nom + (veriler_reel if veriler_reel else [])
    v_min, v_max = min(0, min(tum_veriler)), max(tum_veriler)
    if v_max == v_min: v_max = v_min + 1

    kullan_log = log_olcek and v_max > 0
    if kullan_log: v_min_log = 0; v_max_log = math.log10(v_max) if v_max > 1 else 1

    for i in range(7):
        y = pad_top + ih - int(i / 6 * ih)
        if kullan_log:
            val_log = v_min_log + i / 6 * (v_max_log - v_min_log)
            val = 10 ** val_log
            eksen_metin = f"{val:,.0f} {birim} (L)"
        else:
            val = v_min + i / 6 * (v_max - v_min)
            eksen_metin = f"{val:,.0f} {birim}"
        canvas.create_line(pad_left, y, pad_left + iw, y, fill=r["BORDER"], dash=(4, 4))
        canvas.create_text(pad_left - 8, y, anchor="e", text=eksen_metin, fill=r["TEXT_MUTED"],
                           font=("Courier New", 10))

    n = len(veriler_nom)
    adim = max(1, n // 10)
    for i in range(0, n, adim):
        x = pad_left + int(i / (n - 1) * iw) if n > 1 else pad_left
        canvas.create_line(x, pad_top, x, pad_top + ih, fill=r["BORDER"], dash=(4, 4))
        canvas.create_text(x, pad_top + ih + 15, anchor="n", text=str(i), fill=r["TEXT_MUTED"],
                           font=("Courier New", 10))

    alt_metin = "Süre (Dönem)" if not kullan_log else "Süre (Dönem) - Logaritmik Ölçek Aktif"
    canvas.create_text(W // 2, H - 20, text=alt_metin, fill=r["TEXT_MUTED"], font=("Courier New", 11))
    canvas.create_text(W // 2, 20, text=baslik, fill=r["TEXT_PRIMARY"], font=("Courier New", 13, "bold"))

    if n < 2: return

    def veri2pix(idx, val):
        px = pad_left + int(idx / (n - 1) * iw)
        if kullan_log:
            val_l = math.log10(val) if val > 1 else 0;
            py = pad_top + ih - int(
                (val_l - v_min_log) / (v_max_log - v_min_log) * ih)
        else:
            py = pad_top + ih - int((val - v_min) / (v_max - v_min) * ih)
        return px, py

    def cizgi_olustur(veriler, renk, kalinlik, is_dashed=False):
        if not veriler: return
        noktalar = []
        for i, v in enumerate(veriler):
            px, py = veri2pix(i, v)
            noktalar.extend([px, py])
        if len(noktalar) >= 4: canvas.create_line(noktalar, fill=renk, width=kalinlik, smooth=False,
                                                  dash=(4, 4) if is_dashed else None)
        son_px, son_py = veri2pix(n - 1, veriler[-1])
        canvas.create_oval(son_px - 5, son_py - 5, son_px + 5, son_py + 5, fill=renk, outline=r["PANEL_BG"], width=2)

    if veriler_reel: cizgi_olustur(veriler_reel, r["ACCENT_ORANGE"], 2, True)
    cizgi_olustur(veriler_nom, renk_nom, 3)

    canvas.noktalar = []
    for i, v in enumerate(veriler_nom):
        px, py = veri2pix(i, v)
        canvas.noktalar.append((px, py, v, i))

    tooltip_bg = canvas.create_rectangle(0, 0, 0, 0, fill=r["PANEL_BG"], outline=r["ACCENT_BLUE"], width=2,
                                         state="hidden", tags="tooltip_bg")
    tooltip_text = canvas.create_text(0, 0, text="", fill=r["TEXT_PRIMARY"], font=("Courier New", 11, "bold"),
                                      state="hidden", tags="tooltip_text")
    vurgulu_nokta = [None]

    def on_hover(event):
        x, y = event.x, event.y
        canvas.delete("crosshair")
        if x < pad_left or x > pad_left + iw: on_leave(event); return
        idx = int(round((x - pad_left) / iw * (n - 1))) if n > 1 else 0
        idx = max(0, min(idx, len(veriler_nom) - 1))
        val_nom = veriler_nom[idx]
        px_nom, py_nom = veri2pix(idx, val_nom)
        canvas.create_line(px_nom, pad_top + ih, px_nom, pad_top, dash=(2, 2), fill=r["TEXT_MUTED"], tags="crosshair")
        canvas.create_oval(px_nom - 5, py_nom - 5, px_nom + 5, py_nom + 5, fill=renk_nom, outline=r["PANEL_BG"],
                           width=2, tags="crosshair")
        metin = f"Dönem: {idx}\nNominal: ₺{val_nom:,.2f}"
        if veriler_reel:
            val_reel = veriler_reel[idx]
            _, py_reel = veri2pix(idx, val_reel)
            metin += f"\nReel: ₺{val_reel:,.2f}"
            canvas.create_oval(px_nom - 5, py_reel - 5, px_nom + 5, py_reel + 5, fill=r["ACCENT_ORANGE"],
                               outline=r["PANEL_BG"], width=2, tags="crosshair")
        canvas.itemconfig(tooltip_text, text=metin, state="normal")
        canvas.coords(tooltip_text, px_nom + (-70 if px_nom > W - 150 else 30), py_nom - 20)
        bbox = canvas.bbox(tooltip_text)
        if bbox: canvas.coords(tooltip_bg, bbox[0] - 6, bbox[1] - 6, bbox[2] + 6, bbox[3] + 6); canvas.itemconfig(
            tooltip_bg, state="normal")
        canvas.tag_raise("crosshair")
        canvas.tag_raise("tooltip_bg")
        canvas.tag_raise("tooltip_text")
        if tree is not None and idx > 0:
            cocuklar = tree.get_children()
            if cocuklar and idx <= len(cocuklar):
                hedef = cocuklar[idx - 1]
                if vurgulu_nokta[0] != hedef: vurgulu_nokta[0] = hedef; tree.selection_set(hedef); tree.see(hedef)

    def on_leave(event):
        canvas.delete("crosshair")
        canvas.itemconfig(tooltip_text, state="hidden")
        canvas.itemconfig(tooltip_bg, state="hidden")
        if tree is not None: tree.selection_remove(tree.selection())
        vurgulu_nokta[0] = None

    canvas.bind("<Motion>", on_hover)
    canvas.bind("<Leave>", on_leave)


# ══════════════════════════════════════════════════════════════════════
#  GUI SINIFI
# ══════════════════════════════════════════════════════════════════════
class UygulamaGUI:
    def __init__(self, kok: tk.Tk):
        self.kok = kok
        kok.title("Tasarruf Planı ve Borç Azaltma Modeli: Dizilerin Yakınsaklığı")

        kok.update_idletasks()
        ekr_w = kok.winfo_screenwidth()
        ekr_h = kok.winfo_screenheight()

        pen_w = max(700, min(int(ekr_w * 0.8), 1400))
        pen_h = max(500, min(int(ekr_h * 0.8), 850))

        x = max(0, (ekr_w - pen_w) // 2)
        y = max(0, (ekr_h - pen_h) // 2)

        kok.geometry(f"{pen_w}x{pen_h}+{x}+{y}")
        kok.minsize(700, 500)

        self.b_yeniden_yapilandir_var = tk.BooleanVar(value=False)

        self.kredi_durumlari = {
            "İhtiyaç Kredisi": {"teminat": "", "baslangic": "", "faiz": "", "vergi": "30", "odeme": "",
                                "vade_var": False, "vade": "", "ara_odemeler": [], "hesaplandi": False,
                                "yeniden": False},
            "Taşıt Kredisi": {"teminat": "", "baslangic": "", "faiz": "", "vergi": "30", "odeme": "", "vade_var": False,
                              "vade": "", "ara_odemeler": [], "hesaplandi": False, "yeniden": False},
            "Konut Kredisi": {"teminat": "", "baslangic": "", "faiz": "", "vergi": "0", "odeme": "", "vade_var": False,
                              "vade": "", "ara_odemeler": [], "hesaplandi": False, "yeniden": False},
            "Özel Kredi": {"teminat": "", "baslangic": "", "faiz": "", "vergi": "", "odeme": "", "vade_var": False,
                           "vade": "", "ara_odemeler": [], "hesaplandi": False, "yeniden": False}
        }

        bugun = datetime.date.today().strftime("%d.%m.%Y")
        self.m_tarih = tk.StringVar(value=bugun)
        self.m_baslangic = tk.StringVar(value="")
        self.m_faiz = tk.StringVar(value="")
        self.m_stopaj = tk.StringVar(value="")
        self.m_enflasyon = tk.StringVar(value="")
        self.m_enflasyon_aktif = tk.BooleanVar(value=False)
        self.m_vade_gun = tk.StringVar(value="")
        self.m_erken_cekim_var = tk.BooleanVar(value=False)
        self.m_duzenli_islem_var = tk.BooleanVar(value=False)
        self.m_yatirim = tk.StringVar(value="")
        self.m_sure = tk.StringVar(value="")
        self.m_erken_cekim_liste = []
        self.m_ozel_islemler_liste = []

        self.b_tarih = tk.StringVar(value=bugun)
        self.b_kredi_tipi = tk.StringVar(value="İhtiyaç Kredisi")
        self.b_teminat = tk.StringVar(value="")
        self.b_baslangic = tk.StringVar(value="")
        self.b_faiz = tk.StringVar(value="")
        self.b_vergi = tk.StringVar(value="30")
        self.b_odeme = tk.StringVar(value="")
        self.b_vade_var = tk.BooleanVar(value=False)
        self.b_vade = tk.StringVar(value="")
        self.b_ara_odemeler_liste = []

        self.karanlik_mod = False
        self.renkler = TEMA_AYDINLIK
        self.m_hesaplandi = self.b_hesaplandi = False
        self._init_tamamlandi = False

        self.ana_cerceve = tk.Frame(self.kok)
        self.ana_cerceve.pack(fill="both", expand=True)
        self._tema_uygula()

        self.b_teminat.trace_add("write", self._bddk_guncelle)
        self.b_baslangic.trace_add("write", self._bddk_guncelle)
        self.b_kredi_tipi.trace_add("write", self._bddk_guncelle)

        self.b_baslangic.trace_add("write", self._dinamik_min_odeme)
        self.b_faiz.trace_add("write", self._dinamik_min_odeme)
        self.b_vergi.trace_add("write", self._dinamik_min_odeme)
        self.b_vade.trace_add("write", self._dinamik_min_odeme)
        self._dinamik_min_odeme()

        self.m_vade_gun.trace_add("write", self._oto_stopaj_guncelle)

        self.kok.bind('<Return>', self._enter_basildi)
        self.kok.bind('<Configure>', self._pencere_boyutu_degisti)
        self.kok.bind_all('<Key>', self._virgul_nokta_cevir, add='+')
        self._init_tamamlandi = True

    def _fare_tekerlegi_bagla(self, canvas, scrollbar):
        def _on_mousewheel(e):
            if scrollbar.winfo_ismapped():
                if hasattr(e, 'delta') and e.delta:
                    canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

        def _on_button4(e):
            if scrollbar.winfo_ismapped(): canvas.yview_scroll(-1, "units")

        def _on_button5(e):
            if scrollbar.winfo_ismapped(): canvas.yview_scroll(1, "units")

        def _bind(e):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
            canvas.bind_all("<Button-4>", _on_button4)
            canvas.bind_all("<Button-5>", _on_button5)

        def _unbind(e):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")

        canvas.bind("<Enter>", _bind)
        canvas.bind("<Leave>", _unbind)

    def _scrollable_frame_olustur(self, parent_frame):
        r = self.renkler
        container = tk.Frame(parent_frame, bg=r["PANEL_BG"])
        canvas = tk.Canvas(container, bg=r["PANEL_BG"], highlightthickness=0, height=1)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        icerik_frame = tk.Frame(canvas, bg=r["PANEL_BG"])

        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)

        canvas_window = canvas.create_window((0, 0), window=icerik_frame, anchor="nw")

        def _guncelle(event=None):
            self.kok.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
            child_count = len([w for w in icerik_frame.winfo_children() if w.winfo_exists() and w.winfo_ismapped()])
            row_h = 35

            if child_count == 0:
                canvas.configure(height=1)
                scrollbar.pack_forget()
            elif child_count <= 2:
                canvas.configure(height=child_count * row_h)
                scrollbar.pack_forget()
            else:
                canvas.configure(height=2 * row_h)
                scrollbar.pack(side="right", fill="y")

        icerik_frame.bind("<Configure>", _guncelle)
        icerik_frame.guncelle = _guncelle

        def _on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", _on_canvas_configure)
        self._fare_tekerlegi_bagla(canvas, scrollbar)

        return container, icerik_frame

    def _standart_scrollable_frame(self, parent_frame, bg_color, padx=0, pady=0):
        container = tk.Frame(parent_frame, bg=self.renkler["DARK_BG"])

        outer_panel = tk.Frame(container, bg=bg_color)
        outer_panel.pack(fill="both", expand=True)

        canvas = tk.Canvas(outer_panel, bg=bg_color, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer_panel, orient="vertical", command=canvas.yview)

        icerik_frame = tk.Frame(canvas, bg=bg_color, padx=padx, pady=pady)

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        canvas_window = canvas.create_window((0, 0), window=icerik_frame, anchor="nw")

        def _on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        icerik_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind("<Configure>", _on_canvas_configure)
        self._fare_tekerlegi_bagla(canvas, scrollbar)

        return container, icerik_frame

    def _sonuclari_yazdir(self, parent_frame, veriler, renk):
        for w in parent_frame.winfo_children():
            w.destroy()

        r = self.renkler
        parent_frame.grid_columnconfigure(0, weight=0)
        parent_frame.grid_columnconfigure(1, weight=1)

        for i, (baslik, deger) in enumerate(veriler):
            if baslik == "" and deger == "":
                tk.Frame(parent_frame, height=6, bg=r["PANEL_BG"]).grid(row=i, column=0, columnspan=2)
            elif deger == "":
                lbl = tk.Label(parent_frame, text=baslik, bg=r["PANEL_BG"], fg=r["TEXT_MUTED"],
                               font=("Courier New", 10, "italic"), anchor="nw", justify="left")
                lbl.grid(row=i, column=0, columnspan=2, sticky="we", pady=2)
                lbl.bind("<Configure>", lambda e, l=lbl: l.config(wraplength=max(50, e.width - 5)))
            else:
                lbl_b = tk.Label(parent_frame, text=baslik, bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                                 font=("Courier New", 11, "bold"), anchor="nw", justify="left")
                lbl_b.grid(row=i, column=0, sticky="nw", padx=(0, 6), pady=1)

                lbl_d = tk.Label(parent_frame, text=deger, bg=r["PANEL_BG"], fg=renk, font=("Courier New", 11),
                                 anchor="nw", justify="left")
                lbl_d.grid(row=i, column=1, sticky="nwe", pady=1)
                lbl_d.bind("<Configure>", lambda e, l=lbl_d: l.config(wraplength=max(50, e.width - 5)))

    def _bddk_sinirlari_hesapla(self):
        tip = self.b_kredi_tipi.get()
        teminat_str = self.b_teminat.get().replace('.', '').replace(',', '.')
        borc_str = self.b_baslangic.get().replace('.', '').replace(',', '.')
        teminat = float(teminat_str) if teminat_str not in ["", ".", "-"] else 0.0
        borc = float(borc_str) if borc_str not in ["", ".", "-"] else 0.0
        maks_kredi = float('inf')
        maks_vade = 1200
        mesaj = ""
        renk = self.renkler["ACCENT_BLUE"]

        if tip == "İhtiyaç Kredisi":
            if borc > 0:
                if borc <= 125000:
                    maks_vade = 36
                elif borc <= 250000:
                    maks_vade = 24
                else:
                    maks_vade = 12
                mesaj = f"BDDK İhtiyaç Limiti: ₺{borc:,.2f} için Maksimum Vade {maks_vade} Ay"
            else:
                mesaj = "BDDK İhtiyaç: 0-125k (36 Ay), 125k-250k (24 Ay), >250k (12 Ay)"

        elif tip == "Taşıt Kredisi":
            if teminat > 0:
                if teminat <= 2500000:
                    orani, maks_vade = 0.70, 48
                elif teminat <= 5000000:
                    orani, maks_vade = 0.50, 36
                elif teminat <= 6500000:
                    orani, maks_vade = 0.30, 24
                elif teminat <= 7500000:
                    orani, maks_vade = 0.20, 12
                else:
                    orani, maks_vade = 0.0, 0
                maks_kredi = teminat * orani
                if maks_vade > 0:
                    mesaj = f"BDDK Taşıt (LTV %{int(orani * 100)}): Çekilebilir Maks. Kredi ₺{maks_kredi:,.2f} | Maks. Vade {maks_vade} Ay"
                else:
                    mesaj = "BDDK Taşıt: 7.5M TL üzeri araçlar için kredi kullanımı yasal olarak engellenmiştir!";
                    renk = self.renkler["ACCENT_RED"]
            else:
                mesaj = "Lütfen önce araç fatura değerini giriniz."

        elif tip == "Konut Kredisi":
            if teminat > 0:
                if teminat <= 5000000:
                    orani = 0.90
                elif teminat <= 10000000:
                    orani = 0.80
                elif teminat <= 20000000:
                    orani = 0.70
                else:
                    orani = 0.50
                maks_vade = 120
                maks_kredi = teminat * orani
                mesaj = f"BDDK Konut (LTV %{int(orani * 100)}): Çekilebilir Maks. Kredi ₺{maks_kredi:,.2f} | Maks. Vade {maks_vade} Ay"
            else:
                mesaj = "Lütfen önce konut ekspertiz değerini giriniz."
        else:
            mesaj = "Özel Kredi: BDDK limit kısıtlaması uygulanmaz."

        return maks_kredi, maks_vade, mesaj, renk

    def _bddk_guncelle(self, *args):
        try:
            maks_kredi, maks_vade, mesaj, renk = self._bddk_sinirlari_hesapla()
            self.lbl_bddk_bilgi.config(text=mesaj, fg=renk)
        except Exception:
            pass

    def _oto_stopaj_guncelle(self, *args):
        try:
            vade = int(self.m_vade_gun.get())
            oran = stopaj_hesapla(vade)
            s = f"{float(oran):g}"
            self.m_stopaj.set(s)
        except (ValueError, Exception):
            self.m_stopaj.set("")

    def _dinamik_yazdir(self, lbl, metin, base_size, is_bold=False):
        satir_sayisi = metin.count('\n') + sum(len(satir) // 45 for satir in metin.split('\n'))
        if satir_sayisi > 8:
            size = base_size - 2
        elif satir_sayisi > 5:
            size = base_size - 1
        else:
            size = base_size
        weight = "bold" if is_bold else "normal"
        lbl.config(text=metin, font=("Courier New", max(8, size), weight))

    def _ikon_olustur(self, parent, tooltip_text, underline_words=None):
        icon_lbl = tk.Label(parent, text="ⓘ", fg=self.renkler["ACCENT_BLUE"], bg=self.renkler["PANEL_BG"],
                            font=("Segoe UI Symbol", 10, "bold"), cursor="question_arrow")
        t = Tooltip(icon_lbl, tooltip_text, self.renkler)
        if underline_words: t.underline_words = underline_words
        icon_lbl.tooltip = t
        return icon_lbl

    def _virgul_nokta_cevir(self, event):
        if event.char == ',' and isinstance(event.widget, tk.Entry) and not getattr(event.widget,
                                                                                    'is_currency_formatted', False):
            try:
                event.widget.insert(event.widget.index(tk.INSERT), '.');
                return 'break'
            except:
                pass

    def _para_formatla_event(self, event):
        if event.keysym in ('Left', 'Right', 'Up', 'Down', 'Home', 'End', 'Shift_L', 'Shift_R'): return
        w = event.widget
        raw_val = w.get()
        if not raw_val: return
        c_before = len(raw_val[:w.index(tk.INSERT)].replace('.', ''))
        parts = raw_val.replace('.', '').split(',')
        int_part = ''.join(c for c in parts[0] if c.isdigit())
        f_int = f"{int(int_part):,}".replace(',', '.') if int_part else ("0" if len(parts) > 1 else "")
        new_val = f"{f_int},{''.join(c for c in parts[1] if c.isdigit())}" if len(parts) > 1 else (
            f"{f_int}," if ',' in raw_val else f_int)
        if raw_val != new_val:
            w.delete(0, tk.END)
            w.insert(0, new_val)
            non_dot = 0
            for i, c in enumerate(new_val):
                if non_dot == c_before: w.icursor(i); break
                if c != '.': non_dot += 1
            else:
                w.icursor(len(new_val))

    def _pencere_boyutu_degisti(self, event):
        if event.widget == self.kok: self.kok.after(50, self._grafikleri_yenile)

    def _sekme_degisti(self, event):
        self.kok.after(50, self._grafikleri_yenile)

    def _notebook_tiklandi(self, event):
        try:
            if self.nb.identify(event.x, event.y):
                self.kok.after(50, self._gizli_parametre_kontrol)
        except Exception:
            pass

    def _gizli_parametre_kontrol(self):
        try:
            secili = self.nb.index(self.nb.select())
            if secili == 0 and not self.m_giris_acik:
                self._m_toggle_giris()
            elif secili == 1 and not self.b_giris_acik:
                self._b_toggle_giris()
        except Exception:
            pass

    def _grafikleri_yenile(self):
        self.kok.update_idletasks()
        self.m_canvas.update_idletasks()
        self.b_canvas.update_idletasks()
        if self.m_hesaplandi:
            self._mevduat_hesapla(sessiz=True)
        else:
            self._bosh_grafik_ciz(self.m_canvas, self.renkler, "Tasarruf Eğrisi")
        if self.b_hesaplandi:
            self._borc_hesapla(sessiz=True)
        else:
            self._bosh_grafik_ciz(self.b_canvas, self.renkler, "Borç Eğrisi")

    def _ilk_cizimleri_yap(self):
        self.kok.update_idletasks()
        if not self.m_hesaplandi: self._bosh_grafik_ciz(self.m_canvas, self.renkler, "Tasarruf Eğrisi")
        if not self.b_hesaplandi: self._bosh_grafik_ciz(self.b_canvas, self.renkler, "Borç Eğrisi")

    def _bosh_grafik_ciz(self, canvas, r, baslik):
        canvas.delete("all")
        canvas.unbind("<Motion>")
        canvas.unbind("<Leave>")
        canvas.noktalar = []

        W = canvas.winfo_width() if canvas.winfo_width() > 50 else self.kok.winfo_width() // 2
        H = canvas.winfo_height() if canvas.winfo_height() > 50 else 300
        pad_left, pad_right, pad_top, pad_bottom = 85, 30, 50, 60
        iw, ih = W - pad_left - pad_right, H - pad_top - pad_bottom
        canvas.configure(bg=r["CHART_BG"])
        canvas.create_rectangle(pad_left, pad_top, pad_left + iw, pad_top + ih, fill=r["PANEL_BG"], outline=r["BORDER"])
        canvas.create_text(pad_left + (iw // 2), pad_top + (ih // 2),
                           text="Grafiği görmek için 'Hesapla' butonuna basınız.", fill=r["TEXT_MUTED"],
                           font=("Courier New", 12, "italic"), justify="center", width=max(iw - 20, 10))
        canvas.create_text(W // 2, 20, text=baslik, fill=r["TEXT_PRIMARY"], font=("Courier New", 13, "bold"))

    def _genel_dogrulama(self, P, tip):
        if P == "": return True
        if tip == "yuzde": return (P == ".") or (
                P.count('.') <= 1 and '-' not in P and all(c in "0123456789." for c in P) and (
                '.' not in P or len(P.split('.')[1]) <= 2) and len(P.split('.')[0]) <= 3)
        if tip == "ondalik": return P in ["-", ".", "-."] or (
                P.count('.') <= 1 and P.count('-') <= 1 and (P.find('-') in [-1, 0]) and all(
            c in "0123456789.-" for c in P) and ('.' not in P or len(P.split('.')[1]) <= 2))
        if tip == "ondalik_pozitif": return (P == ".") or (
                P.count('.') <= 1 and '-' not in P and all(c in "0123456789." for c in P) and (
                '.' not in P or len(P.split('.')[1]) <= 2))
        if tip == "tamsayi": return all(c in "0123456789" for c in P)
        if tip == "tarih": return len(P) <= 10 and all(c in "0123456789." for c in P)
        if tip == "para_gorsel":
            c = P.replace('.', '')
            if c.count(',') <= 1 and all(ch in "0123456789," for ch in c):
                parts = c.split(',')
                if len(parts[0]) <= 10 and (len(parts) == 1 or len(parts[1]) <= 2):
                    return True
            return False
        return True

    def _enter_basildi(self, event):
        if self.nb.index(self.nb.select()) == 0:
            self._mevduat_hesapla(sessiz=False)
        else:
            self._borc_hesapla(sessiz=False)

    def _dinamik_min_odeme(self, *args):
        if not hasattr(self, 'lbl_dinamik_min'): return
        try:
            b_bas_str = self.b_baslangic.get().replace('.', '').replace(',', '.')
            B0 = Decimal(b_bas_str) if b_bas_str not in ["", ".", "-"] else Decimal('0')

            faiz_str = self.b_faiz.get()
            vergi_str = self.b_vergi.get()
            F = Decimal(faiz_str) / Decimal('100') if faiz_str not in ["", ".", "-"] else Decimal('0')
            vergi_orani_d = Decimal(vergi_str) / Decimal('100') if vergi_str not in ["", ".", "-"] else Decimal('0')
            r_aylik = (F / Decimal('12')) * (Decimal('1') + vergi_orani_d)

            tahakkuk = (B0 * r_aylik).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)
            maks_kredi, maks_vade, mesaj, renk = self._bddk_sinirlari_hesapla()
            tip = self.b_kredi_tipi.get()

            if tip != "Özel Kredi" and maks_vade > 0 and maks_vade < 1200:
                if r_aylik > 0:
                    yasal_min_odeme = (B0 * (r_aylik / (1 - (1 + r_aylik) ** (-maks_vade)))).quantize(Decimal('.01'),
                                                                                                      rounding=ROUND_HALF_UP)
                else:
                    yasal_min_odeme = (B0 / Decimal(maks_vade)).quantize(Decimal('.01'),
                                                                         rounding=ROUND_HALF_UP) if maks_vade > 0 else Decimal(
                        '0')
                self.lbl_dinamik_min.config(
                    text=f"(Yasal sınır, {maks_vade} ayı karşılamak için gereken minimum ödeme tutarı: ₺{yasal_min_odeme:,.2f})")
            elif maks_vade == 0:
                self.lbl_dinamik_min.config(text="(Kredi kullanımı yasal olarak engellenmiştir.)")
            else:
                self.lbl_dinamik_min.config(text=f"Tahakkuk Eden Faiz (Min. Ödeme): ₺{tahakkuk:,.2f}")
        except:
            self.lbl_dinamik_min.config(text="Bekleniyor...")

        if not hasattr(self, 'lbl_vade_min_odeme'): return
        try:
            if not self.b_vade_var.get() or self.b_vade.get() == "":
                self.lbl_vade_min_odeme.config(text="")
                return

            vade_degeri = int(self.b_vade.get())
            if vade_degeri < 1: raise ValueError

            maks_kredi, maks_vade, mesaj, renk = self._bddk_sinirlari_hesapla()

            if maks_vade != float('inf') and maks_vade > 0 and vade_degeri > maks_vade:
                vade_degeri = maks_vade
                self.kok.after(10, lambda: self.b_vade.set(str(maks_vade)))

            B0 = Decimal(self.b_baslangic.get().replace('.', '').replace(',', '.')) if self.b_baslangic.get() not in [
                "", ".", "-"] else Decimal('0')
            F = Decimal(self.b_faiz.get()) / Decimal('100') if self.b_faiz.get() not in ["", ".", "-"] else Decimal('0')
            vergi_orani_d = Decimal(self.b_vergi.get()) / Decimal('100') if self.b_vergi.get() not in ["", ".",
                                                                                                       "-"] else Decimal(
                '0')
            r_aylik = (F / Decimal('12')) * (Decimal('1') + vergi_orani_d)

            if r_aylik > 0:
                min_odeme_vade = (B0 * (r_aylik / (1 - (1 + r_aylik) ** (-vade_degeri)))).quantize(Decimal('.01'),
                                                                                                   rounding=ROUND_HALF_UP)
            else:
                min_odeme_vade = (B0 / Decimal(vade_degeri)).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)

            self.lbl_vade_min_odeme.config(
                text=f"Vade Hedefi ({vade_degeri} ay) Taksit Tutarı: ₺{min_odeme_vade:,.2f}")
            if self.b_vade_var.get():
                self.b_odeme.set(f"{min_odeme_vade:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        except:
            self.lbl_vade_min_odeme.config(text="")

    def _tema_degistir(self):
        if hasattr(self, 'kredi_durumlari') and getattr(self, '_init_tamamlandi', False):
            eski_tip = self.b_kredi_tipi.get()
            self.kredi_durumlari[eski_tip] = {
                "teminat": self.b_teminat.get(), "baslangic": self.b_baslangic.get(), "faiz": self.b_faiz.get(),
                "vergi": self.b_vergi.get(), "odeme": self.b_odeme.get(), "vade_var": self.b_vade_var.get(),
                "vade": self.b_vade.get(), "ara_odemeler": [(d.get(), t.get()) for d, t in self.b_ara_odemeler_liste],
                "hesaplandi": self.b_hesaplandi, "yeniden": self.b_yeniden_yapilandir_var.get()
            }

        m_erken_kayit = [(k[0].get(), k[1].get(), k[2]) for k in self.m_erken_cekim_liste]
        m_ozel_kayit = [(d.get(), t.get()) for d, t in self.m_ozel_islemler_liste]

        self._init_tamamlandi = False
        aktif_sekme = self.nb.index(self.nb.select())
        self.karanlik_mod = not self.karanlik_mod
        self.renkler = TEMA_KARANLIK if self.karanlik_mod else TEMA_AYDINLIK

        for widget in self.ana_cerceve.winfo_children(): widget.destroy()

        self.m_erken_cekim_liste.clear()
        self.m_ozel_islemler_liste.clear()
        self.b_ara_odemeler_liste.clear()
        self._tema_uygula()
        self.nb.select(aktif_sekme)

        for d, t, tip in m_erken_kayit:
            self._erken_cekim_ekle(d, t, tip=tip)
        for d, t in m_ozel_kayit: self._ozel_islem_ekle(d, t)

        self.kok.update_idletasks()
        self._init_tamamlandi = True

        self.kok.after(100, lambda: self._mevduat_hesapla(sessiz=True) if self.m_hesaplandi else self._bosh_grafik_ciz(
            self.m_canvas, self.renkler, "Tasarruf Eğrisi"))
        self.kok.after(100, lambda: self._borc_hesapla(sessiz=True) if self.b_hesaplandi else self._bosh_grafik_ciz(
            self.b_canvas, self.renkler, "Borç Eğrisi"))

    def _tema_uygula(self):
        r = self.renkler
        self.kok.configure(bg=r["DARK_BG"])
        self.ana_cerceve.configure(bg=r["DARK_BG"])
        stil = ttk.Style()
        stil.theme_use("clam")
        stil.configure("TNotebook", background=r["DARK_BG"], borderwidth=0)
        stil.configure("TNotebook.Tab", background=r["PANEL_BG"], foreground=r["TEXT_MUTED"], padding=[20, 8],
                       font=("Courier New", 12, "bold"), borderwidth=0)
        stil.map("TNotebook.Tab", background=[("selected", r["PANEL_BG"])], foreground=[("selected", r["ACCENT_BLUE"])])
        stil.configure("TFrame", background=r["DARK_BG"])
        stil.configure("Treeview", background=r["ENTRY_BG"], foreground=r["TEXT_PRIMARY"],
                       fieldbackground=r["ENTRY_BG"], rowheight=28, font=("Courier New", 10))
        stil.configure("Treeview.Heading", background=r["PANEL_BG"], foreground=r["ACCENT_BLUE"],
                       font=("Courier New", 10, "bold"))
        stil.map('Treeview', background=[('selected', r["ACCENT_BLUE"])], foreground=[('selected', r["DARK_BG"])])
        self._arayuz_olustur()

    def _arayuz_olustur(self):
        r = self.renkler
        baslik_frame = tk.Frame(self.ana_cerceve, bg=r["DARK_BG"])
        baslik_frame.pack(fill="x", padx=24, pady=(12, 0))
        tk.Label(baslik_frame, text="Tasarruf Planı ve Borç Azaltma Modeli: Dizilerin Yakınsaklığı", bg=r["DARK_BG"],
                 fg=r["ACCENT_BLUE"], font=("Courier New", 15, "bold")).pack(side="left")
        tk.Button(baslik_frame, text="☀️/🌙 Temayı Değiştir", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                  font=("Courier New", 10, "bold"), relief="solid", bd=1, cursor="hand2",
                  command=self._tema_degistir).pack(side="right")

        self.nb = ttk.Notebook(self.ana_cerceve)
        self.nb.pack(fill="both", expand=True, padx=16, pady=8)
        self.nb.bind("<<NotebookTabChanged>>", self._sekme_degisti)
        self.nb.bind("<Button-1>", self._notebook_tiklandi)

        sekme_mevduat = ttk.Frame(self.nb)
        sekme_borc = ttk.Frame(self.nb)
        self.nb.add(sekme_mevduat, text="  🏦  Tasarruf Planı  ")
        self.nb.add(sekme_borc, text="  📉  Borç Ekstresi  ")

        self._mevduat_sekmesi(sekme_mevduat)
        self._borc_sekmesi(sekme_borc)
        self._dinamik_min_odeme()
        self.kok.after(200, self._ilk_cizimleri_yap)

    def _tarih_entry_olustur(self, parent, tarih_var, satir):
        r = self.renkler

        vcmd = (self.kok.register(lambda P: self._genel_dogrulama(P, "tarih")), '%P')

        comp_frame = tk.Frame(parent, bg=r["PANEL_BG"])
        comp_frame.grid(row=satir, column=1, sticky="w", pady=2)

        entry = CustomEntry(comp_frame, is_dark=self.karanlik_mod, textvariable=tarih_var, width=12,
                            bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                            insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 12),
                            validate="key", validatecommand=vcmd)
        entry.pack(side="left")

        def _otomatik_nokta(event):
            if event.keysym in ('BackSpace', 'Delete', 'Left', 'Right', 'Home', 'End'): return
            deger = tarih_var.get()
            if len(deger) == 2 and deger.count('.') == 0:
                tarih_var.set(deger + '.');
                entry.icursor(3)
            elif len(deger) == 5 and deger.count('.') == 1:
                tarih_var.set(deger + '.');
                entry.icursor(6)

        entry.bind('<KeyRelease>', _otomatik_nokta)
        tk.Button(comp_frame, text="📅", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], relief="flat", font=("Courier New", 11),
                  cursor="hand2", command=lambda: self._takvim_ac(tarih_var, entry)).pack(side="left", padx=(2, 0))
        return entry

    def _takvim_ac(self, tarih_var, entry_widget):
        if getattr(self, "aktif_takvim", None) and self.aktif_takvim.winfo_exists():
            eski_var = self.aktif_takvim.tarih_var
            self.aktif_takvim.destroy()
            self.aktif_takvim = None
            if eski_var == tarih_var:
                return

        self.aktif_takvim = TakvimPopup(self.kok, tarih_var, self.renkler, entry_widget)

    def _erken_cekim_toggle(self):
        pass

    def _erken_cekim_ekle(self, d_val="", t_val="", tip="yatirma"):
        r = self.renkler
        f = tk.Frame(self.erken_cekim_icerik, bg=r["PANEL_BG"])
        f.pack(fill="x", pady=1)

        if tip == "cekme":
            tarih_var = tk.StringVar(value=str(d_val) if d_val else "")
            t_var = tk.StringVar(value=str(t_val))

            tk.Label(f, text="Çekim Tarihi:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                     font=("Courier New", 10, "bold")).pack(side="left")

            vcmd_tarih = (self.kok.register(lambda P: self._genel_dogrulama(P, "tarih")), '%P')
            cekim_tarih_bg = r.get("TEXTBOX_BG", r["ENTRY_BG"])
            cekim_tarih_fg = r.get("TEXTBOX_FG", r["TEXT_PRIMARY"])
            tarih_entry = CustomEntry(f, is_dark=self.karanlik_mod, textvariable=tarih_var, width=12,
                                      bg=cekim_tarih_bg, fg=cekim_tarih_fg,
                                      insertbackground=cekim_tarih_fg,
                                      font=("Courier New", 10), validate="key", validatecommand=vcmd_tarih)
            tarih_entry.pack(side="left", padx=(2, 2))

            def _otomatik_nokta_cekim(event, tv=tarih_var, widget=tarih_entry):
                if event.keysym in ('BackSpace', 'Delete', 'Left', 'Right', 'Home', 'End'): return
                deger = tv.get()
                if len(deger) == 2 and deger.count('.') == 0:
                    tv.set(deger + '.')
                    widget.icursor(3)
                elif len(deger) == 5 and deger.count('.') == 1:
                    tv.set(deger + '.')
                    widget.icursor(6)

            tarih_entry.bind('<KeyRelease>', _otomatik_nokta_cekim)

            def _takvim_ac_cekim(entry=tarih_entry, tv=tarih_var):
                if getattr(self, "aktif_takvim", None) and self.aktif_takvim.winfo_exists():
                    eski_var = self.aktif_takvim.tarih_var
                    self.aktif_takvim.destroy()
                    self.aktif_takvim = None
                    if eski_var == tv:
                        return
                try:
                    min_t = datetime.datetime.strptime(self.m_tarih.get(), "%d.%m.%Y").date()
                except Exception:
                    min_t = None
                self.aktif_takvim = TakvimPopup(self.kok, tv, self.renkler, entry, min_tarih=min_t)

            takvim_btn = tk.Button(f, text="📅", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], relief="flat",
                                   font=("Courier New", 10), cursor="hand2",
                                   command=_takvim_ac_cekim)
            takvim_btn.pack(side="left", padx=(0, 6))
            _takvim_tooltip_text = "Para çektiğiniz tarihte vadeniz bozulur.\nMevduatın yeni başlangıç tarihi bu tarih olacaktır."
            Tooltip(takvim_btn, _takvim_tooltip_text, r)

            tk.Label(f, text="Tutar (₺):", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                     font=("Courier New", 10, "bold")).pack(side="left")

            vcmd_dec_pozitif = (self.kok.register(lambda P: self._genel_dogrulama(P, "ondalik_pozitif")), '%P')
            CustomEntry(f, is_dark=self.karanlik_mod, textvariable=t_var, width=10,
                        bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                        insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 10),
                        validate="key", validatecommand=vcmd_dec_pozitif).pack(side="left", padx=(2, 6))

            tk.Button(f, text="✕", bg=r["ACCENT_RED"], fg="#FFFFFF", relief="flat", font=("Courier New", 9, "bold"),
                      cursor="hand2",
                      command=lambda frm=f, v=(tarih_var, t_var, "cekme"): self._erken_cekim_sil(frm, v)).pack(
                side="left")
            self.m_erken_cekim_liste.append((tarih_var, t_var, "cekme"))

        else:
            d_var = tk.StringVar(value=str(d_val))
            t_var = tk.StringVar(value=str(t_val))
            vcmd_int = (self.kok.register(lambda P: self._genel_dogrulama(P, "tamsayi")), '%P')
            vcmd_dec = (self.kok.register(lambda P: self._genel_dogrulama(P, "ondalik_pozitif")), '%P')
            tk.Label(f, text="Dönem:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], font=("Courier New", 10)).pack(
                side="left")

            CustomEntry(f, is_dark=self.karanlik_mod, textvariable=d_var, width=5,
                        bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                        insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 11),
                        validate="key", validatecommand=vcmd_int).pack(side="left", padx=(2, 6))

            tk.Label(f, text="Tutar (₺):", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], font=("Courier New", 10)).pack(
                side="left")

            CustomEntry(f, is_dark=self.karanlik_mod, textvariable=t_var, width=10,
                        bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                        insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 11),
                        validate="key", validatecommand=vcmd_dec).pack(side="left", padx=(2, 6))

            tk.Button(f, text="✕", bg=r["ACCENT_RED"], fg="#FFFFFF", relief="flat", font=("Courier New", 9, "bold"),
                      cursor="hand2",
                      command=lambda frm=f, v=(d_var, t_var, "yatirma"): self._erken_cekim_sil(frm, v)).pack(
                side="left")
            self.m_erken_cekim_liste.append((d_var, t_var, "yatirma"))

    def _erken_cekim_sil(self, frame, var_tuple):
        if var_tuple in self.m_erken_cekim_liste: self.m_erken_cekim_liste.remove(var_tuple)
        frame.destroy()
        self.erken_cekim_icerik.guncelle()

    def _duzenli_islem_toggle(self):
        if self.m_duzenli_islem_var.get():
            self.m_yatirim_frame.grid(row=8, column=1, sticky="w", pady=4)
        else:
            self.m_yatirim_frame.grid_remove();
            self.m_yatirim.set("")

    def _enflasyon_toggle(self):
        if self.m_enflasyon_aktif.get():
            self.enf_entry_frame.grid(row=7, column=1, sticky="w", pady=4)
        else:
            self.enf_entry_frame.grid_remove();
            self.m_enflasyon.set("")

    def _b_yapi_guncelle(self):
        if len(self.b_ara_odemeler_liste) > 0:
            self.b_yapi_frame.grid(row=11, column=0, columnspan=3, sticky="w", pady=4)
        else:
            self.b_yapi_frame.grid_remove()
            self.b_yeniden_yapilandir_var.set(False)

    def _b_vade_toggle(self):
        if self.b_vade_var.get():
            self.b_vade_entry_frame.grid(row=10, column=1, sticky="w", pady=4)
            self.entry_b_odeme.config(state="disabled")
            self._dinamik_min_odeme()
        else:
            self.b_vade_entry_frame.grid_remove()
            self.b_vade.set("")
            self.entry_b_odeme.config(state="normal")

    def _etiket_giris(self, parent, metin, attr, satir, v_tipi, tooltip=""):
        r = self.renkler
        tk.Label(parent, text=metin, bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], font=("Courier New", 11), width=28,
                 anchor="w").grid(row=satir,
                                  column=0,
                                  sticky="w",
                                  pady=2,
                                  padx=(0, 8))
        vcmd = (self.kok.register(lambda P: self._genel_dogrulama(P, v_tipi)), '%P')
        entry = CustomEntry(parent, is_dark=self.karanlik_mod, textvariable=getattr(self, attr), width=15,
                            bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                            insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 12),
                            validate="key", validatecommand=vcmd)
        entry.grid(row=satir, column=1, sticky="w", pady=2)
        if tooltip: self._ikon_olustur(parent, tooltip).grid(row=satir, column=2, sticky="w", padx=(4, 0))
        return entry

    def _mevduat_sekmesi(self, parent):
        r = self.renkler
        icerik = tk.Frame(parent, bg=r["DARK_BG"])
        icerik.pack(fill="both", expand=True, padx=8, pady=4)
        icerik.grid_columnconfigure(1, weight=1)
        icerik.grid_rowconfigure(0, weight=1)

        sol_panel = tk.Frame(icerik, bg=r["DARK_BG"])
        sol_panel.grid(row=0, column=0, sticky="ns", padx=(0, 16))
        sag_panel = tk.Frame(icerik, bg=r["DARK_BG"])
        sag_panel.grid(row=0, column=1, sticky="nsew")

        giris = tk.Frame(sol_panel, bg=r["PANEL_BG"], padx=16, pady=6)
        giris.pack(fill="x", pady=(0, 8))

        self.m_giris_acik = True

        def _m_toggle_giris():
            if self.m_giris_acik:
                giris.pack_forget()
                self.m_toggle_btn.config(text="▼  Parametreler", bg=r["ACCENT_DARK_BLUE"], fg="#FFFFFF")
                self.m_giris_acik = False
            else:
                giris.pack(fill="x", pady=(0, 8), before=self.m_toggle_btn_frame)
                self.m_toggle_btn.config(text="▲  Parametreleri Gizle", bg=r["PANEL_BG"], fg=r["ACCENT_BLUE"])
                self.m_giris_acik = True

        self.m_toggle_btn_frame = tk.Frame(sol_panel, bg=r["DARK_BG"])
        self.m_toggle_btn_frame.pack(fill="x", pady=(0, 4))
        self.m_toggle_btn = tk.Button(self.m_toggle_btn_frame, text="▲  Parametreleri Gizle",
                                      bg=r["PANEL_BG"], fg=r["ACCENT_BLUE"],
                                      font=("Courier New", 10, "bold"), relief="flat", cursor="hand2",
                                      command=_m_toggle_giris, anchor="w")
        self.m_toggle_btn.pack(fill="x")
        self._m_toggle_giris = _m_toggle_giris

        giris.grid_columnconfigure(0, minsize=260, weight=0)
        giris.grid_columnconfigure(1, weight=0)
        giris.grid_columnconfigure(2, weight=1)

        tk.Label(giris, text="Başlangıç Tarihi:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                 font=("Courier New", 11), width=28, anchor="w").grid(row=0, column=0, sticky="w", pady=2, padx=(0, 8))
        self._tarih_entry_olustur(giris, self.m_tarih, 0)

        entry_m_bas = self._etiket_giris(giris, "Başlangıç Bakiye (₺):", "m_baslangic", 1, "para_gorsel",
                                         "Başlangıçtaki anapara tutarını giriniz.\nMaksimum 10 basamaklı bir değer girin.")
        entry_m_bas.is_currency_formatted = True
        entry_m_bas.bind('<KeyRelease>', self._para_formatla_event, add='+')

        self._etiket_giris(giris, "Brüt Faiz Oranı (%):", "m_faiz", 2, "yuzde",
                           "Yıllık brüt banka faiz oranını giriniz.\n%0.0 - %100.0 arasında bir değer girin.")
        self._etiket_giris(giris, "Vade Günü (Tamsayı):", "m_vade_gun", 3, "tamsayi",
                           "Faizin tahakkuk edeceği gün sayısı (Örn: 32).\n1 - 999 arasında gün sayısını girin.")

        self.entry_m_stopaj = self._etiket_giris(giris, "Stopaj (Vergi) (%):", "m_stopaj", 4, "yuzde",
                                                 "Faizden kesilen yasal vergi.\nVade sürenize göre otomatik olarak hesaplanır.")
        self.entry_m_stopaj.config(state="disabled")

        erken_cekim_baslik = tk.Frame(giris, bg=r["PANEL_BG"])
        erken_cekim_baslik.grid(row=5, column=0, sticky="w", pady=(4, 2))
        tk.Label(erken_cekim_baslik, text="Dönem İçi İşlemler:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                 font=("Courier New", 11)).pack(side="left")

        self._ikon_olustur(erken_cekim_baslik,
                           "Dönem bitmeden önce para çekme veya yatırma yapacaksanız belirtiniz.\nPara çekimi vadeyi bozar ve faiz yanar.\nPara yatırma işlemleri vadeyi bozmaz.").pack(
            side="left", padx=(4, 8))

        def _erken_islem_tipi_sec():
            menu = tk.Menu(self.kok, tearoff=0, bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                           activebackground=r["ACCENT_BLUE"], activeforeground=r["DARK_BG"],
                           font=("Courier New", 10))
            menu.add_command(label="Para Yatırma", command=lambda: self._erken_cekim_ekle(tip="yatirma"))
            menu.add_command(label="Para Çekme", command=lambda: self._erken_cekim_ekle(tip="cekme"))
            self.btn_erken_ekle.update_idletasks()
            x = self.btn_erken_ekle.winfo_rootx()
            y = self.btn_erken_ekle.winfo_rooty() + self.btn_erken_ekle.winfo_height()
            menu.tk_popup(x, y)

        self.btn_erken_ekle = tk.Button(giris, text="＋ Ekle", bg=r["ACCENT_BLUE"], fg=r["DARK_BG"],
                                        relief="flat", font=("Courier New", 9, "bold"), cursor="hand2",
                                        command=_erken_islem_tipi_sec)
        self.btn_erken_ekle.grid(row=5, column=1, sticky="e")

        self.erken_cekim_icerik_container, self.erken_cekim_icerik = self._scrollable_frame_olustur(giris)
        self.erken_cekim_icerik_container.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(0, 2))

        self.enf_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        self.enf_frame.grid(row=7, column=0, sticky="w", pady=4)
        tk.Checkbutton(self.enf_frame, text="Reel Alım Gücü (Enflasyon):", variable=self.m_enflasyon_aktif,
                       command=self._enflasyon_toggle, bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                       selectcolor=r["ENTRY_BG"], activebackground=r["PANEL_BG"], activeforeground=r["TEXT_PRIMARY"],
                       font=("Courier New", 11)).pack(side="left", padx=(0, 8))

        self.enf_entry_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        self.enf_entry = CustomEntry(self.enf_entry_frame, is_dark=self.karanlik_mod, textvariable=self.m_enflasyon,
                                     width=8,
                                     bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                                     insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 12),
                                     validate="key",
                                     validatecommand=(self.kok.register(lambda P: self._genel_dogrulama(P, "yuzde")),
                                                      '%P'))
        self.enf_entry.pack(side="left")
        if self.m_enflasyon_aktif.get(): self.enf_entry_frame.grid(row=7, column=1, sticky="w", pady=4)
        enf_icon = self._ikon_olustur(giris,
                                      "Mevcut/beklenen yıllık enflasyon oranını girin.\n0 - 100 arasında bir değer girin.")
        enf_icon.grid(row=7, column=2, sticky="w", padx=(4, 0))

        self.duzenli_islem_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        self.duzenli_islem_frame.grid(row=8, column=0, sticky="w", pady=4)
        tk.Checkbutton(self.duzenli_islem_frame, text="Düzenli İşlem (Her Dönem):", variable=self.m_duzenli_islem_var,
                       command=self._duzenli_islem_toggle, bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                       selectcolor=r["ENTRY_BG"], activebackground=r["PANEL_BG"], activeforeground=r["TEXT_PRIMARY"],
                       font=("Courier New", 11)).pack(side="left", padx=(0, 8))

        self.m_yatirim_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        self.m_yatirim_entry = CustomEntry(self.m_yatirim_frame, is_dark=self.karanlik_mod, textvariable=self.m_yatirim,
                                           width=15,
                                           bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]),
                                           fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                                           insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                                           font=("Courier New", 12), validate="key",
                                           validatecommand=(
                                               self.kok.register(lambda P: self._genel_dogrulama(P, "ondalik")), '%P'))
        self.m_yatirim_entry.pack(side="left")
        if self.m_duzenli_islem_var.get(): self.m_yatirim_frame.grid(row=8, column=1, sticky="w", pady=4)
        self._ikon_olustur(giris, "Vade sonunda her dönem yapacağınız para çekme\nveya yatırma işlemlerini giriniz.",
                           underline_words=["her dönem"]).grid(row=8, column=2, sticky="w", padx=(4, 0))

        ozel_baslik_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        ozel_baslik_frame.grid(row=9, column=0, sticky="w", pady=(4, 2))
        tk.Label(ozel_baslik_frame, text="Özel İşlemler:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                 font=("Courier New", 11)).pack(side="left")
        self._ikon_olustur(ozel_baslik_frame,
                           "Vade sonunda yapacağınız düzensiz para çekme\nve yatırma işlemlerini giriniz.").pack(
            side="left", padx=(4, 0))

        self.btn_ozel_ekle = tk.Button(giris, text="＋ Ekle", bg=r["ACCENT_BLUE"], fg=r["DARK_BG"], relief="flat",
                                       font=("Courier New", 9, "bold"), cursor="hand2", command=self._ozel_islem_ekle)
        self.btn_ozel_ekle.grid(row=9, column=1, sticky="e")

        self.ozel_islem_frame_container, self.ozel_islem_frame = self._scrollable_frame_olustur(giris)
        self.ozel_islem_frame_container.grid(row=10, column=0, columnspan=3, sticky="ew", pady=(0, 2))

        self._etiket_giris(giris, "Tekrar Edecek Dönem (1-120):", "m_sure", 11, "tamsayi",
                           "Simülasyonun toplam uzunluğunu giriniz.\n1 - 120 Dönem arasında bir değer girin.")

        self.btn_frame_m = tk.Frame(giris, bg=r["PANEL_BG"])
        self.btn_frame_m.grid(row=12, column=0, columnspan=3, sticky="ew", pady=(10, 0))
        self.btn_frame_m.grid_columnconfigure(0, weight=4)
        self.btn_frame_m.grid_columnconfigure(1, weight=1)
        self.btn_frame_m.grid_columnconfigure(2, weight=0)

        tk.Button(self.btn_frame_m, text="  Hesapla  ", bg=r["ACCENT_BLUE"], fg="#FFFFFF",
                  font=("Courier New", 11, "bold"),
                  relief="flat", cursor="hand2", command=self._mevduat_hesapla).grid(row=0, column=0, sticky="ew")
        tk.Button(self.btn_frame_m, text="Temizle", bg=r["ACCENT_RED"], fg="#FFFFFF", font=("Courier New", 11, "bold"),
                  relief="flat", cursor="hand2", command=self._mevduat_temizle).grid(row=0, column=1, sticky="ew",
                                                                                     padx=(4, 0))

        self.btn_indir_csv_m = tk.Button(self.btn_frame_m, text="İndir", bg=r["ACCENT_GREEN"], fg="#FFFFFF",
                                         font=("Courier New", 11, "bold"),
                                         relief="flat", cursor="hand2", command=self._indirme_popup_ac_m)
        self.btn_indir_csv_m.grid(row=0, column=2, sticky="ew", padx=(4, 0))
        self.btn_indir_csv_m.grid_remove()

        self.m_sonuc_container, self.m_sonuc_frame = self._standart_scrollable_frame(sol_panel, bg_color=r["PANEL_BG"],
                                                                                     padx=16, pady=8)
        self.m_sonuc_container.pack(fill="both", expand=True, pady=(0, 8))

        self.m_uyari_lbl = tk.Label(self.m_sonuc_frame, text="", bg=r["PANEL_BG"], fg=r["ACCENT_ORANGE"],
                                    font=("Courier New", 11, "bold"), justify="left", anchor="nw")
        self.m_uyari_lbl.pack(anchor="w", pady=(0, 8), fill="x")
        self.m_uyari_lbl.bind("<Configure>", lambda e: e.widget.config(wraplength=max(50, e.width - 5)))

        self.m_sonuc_grid_frame = tk.Frame(self.m_sonuc_frame, bg=r["PANEL_BG"])
        self.m_sonuc_grid_frame.pack(anchor="w", fill="x")

        sag_panel.grid_rowconfigure(0, weight=1)
        sag_panel.grid_rowconfigure(1, weight=2)
        sag_panel.grid_columnconfigure(0, weight=1)
        self.m_canvas = tk.Canvas(sag_panel, bg=r["CHART_BG"], highlightthickness=0)
        self.m_canvas.grid(row=0, column=0, sticky="nsew", pady=(0, 10))

        tablo_frame_m = tk.Frame(sag_panel, bg=r["ENTRY_BG"])
        tablo_frame_m.grid(row=1, column=0, sticky="nsew")
        tablo_frame_m.grid_rowconfigure(1, weight=1)
        tablo_frame_m.grid_rowconfigure(2, weight=0)
        tablo_frame_m.grid_columnconfigure(0, weight=1)

        self.m_tablo_not_lbl = tk.Label(tablo_frame_m, text="", bg=r["ENTRY_BG"], fg=r["ACCENT_ORANGE"],
                                        font=("Courier New", 9, "italic"), justify="left", wraplength=600, anchor="w")
        self.m_tablo_not_lbl.grid(row=0, column=0, columnspan=2, sticky="ew", padx=4, pady=(2, 0))

        self.tree_m = ttk.Treeview(tablo_frame_m, columns=("donem", "bas_tarih", "bit_tarih", "faiz", "islem", "bakiye",
                                                           "reel_bakiye"),
                                   show="headings")
        for c, n, w in [("donem", "Dönem", 50), ("bas_tarih", "Vade Başlangıcı", 105),
                        ("bit_tarih", "Vade Bitişi", 105),
                        ("faiz", "Net Faiz", 110),
                        ("islem", "Nakit Akışı", 110), ("bakiye", "Nominal Bakiye", 130),
                        ("reel_bakiye", "Reel Bakiye", 130)]:
            self.tree_m.heading(c, text=n)
            self.tree_m.column(c, width=w, anchor="center" if c in ("donem", "bas_tarih", "bit_tarih") else "e")
        self.tree_m.tag_configure("evenrow", background=r["ROW_EVEN"])
        self.tree_m.tag_configure("oddrow", background=r["ROW_ODD"])
        self.tree_m.tag_configure("vade_bozuldu", foreground=r["ACCENT_ORANGE"])

        scroll_m = ttk.Scrollbar(tablo_frame_m, orient="vertical", command=self.tree_m.yview)
        scroll_m_x = ttk.Scrollbar(tablo_frame_m, orient="horizontal", command=self.tree_m.xview)
        self.tree_m.configure(yscrollcommand=scroll_m.set, xscrollcommand=scroll_m_x.set)
        self.tree_m.grid(row=1, column=0, sticky="nsew")
        scroll_m.grid(row=1, column=1, sticky="ns")
        scroll_m_x.grid(row=2, column=0, sticky="ew")
        self.tree_m.bind("<<TreeviewSelect>>", lambda e: self._tablo_secim_grafik(self.tree_m, self.m_canvas, "m"))

    def _indirme_popup_ac_m(self):
        if not self.m_hesaplandi:
            self.m_uyari_lbl.config(text="❌ Lütfen önce tasarruf hesaplaması yapınız!", fg=self.renkler["ACCENT_RED"])
            return

        top = tk.Toplevel(self.kok)
        top.title("İndirme Formatı")
        top.geometry("480x150")
        top.configure(bg=self.renkler["PANEL_BG"])
        top.transient(self.kok)
        top.grab_set()

        top.update_idletasks()
        x = self.kok.winfo_rootx() + (self.kok.winfo_width() - top.winfo_width()) // 2
        y = self.kok.winfo_rooty() + (self.kok.winfo_height() - top.winfo_height()) // 2
        top.geometry(f"+{x}+{y}")

        tk.Label(top, text="Hangi formatta indirmek istersiniz?", bg=self.renkler["PANEL_BG"],
                 fg=self.renkler["TEXT_PRIMARY"], font=("Courier New", 11, "bold")).pack(pady=25)

        btn_frame = tk.Frame(top, bg=self.renkler["PANEL_BG"])
        btn_frame.pack(fill="x", padx=20)

        def kaydet_baslat(format_tipi):
            top.destroy()
            if format_tipi == "xlsx":
                dosya = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                     initialfile="Tasarruf Planı",
                                                     filetypes=[("Excel Dosyası", "*.xlsx")],
                                                     title="Excel Olarak Kaydet")
                if dosya:
                    try:
                        basliklar = ["Dönem", "Vade Başlangıcı", "Vade Bitişi", "Net Faiz", "Nakit Akışı",
                                     "Nominal Bakiye", "Reel Bakiye"]
                        satir_verileri = [self.tree_m.item(s)['values'] for s in self.tree_m.get_children()]
                        self._xlsx_zebra_kaydet(dosya, basliklar, satir_verileri,
                                                sayfa_adi="Tasarruf Planı",
                                                tablo_basligi="Tasarruf Planı - Detaylı Ekstre")
                        self.m_uyari_lbl.config(text=f"✅ Zebra tasarruf planı Excel olarak kaydedildi:\n{dosya}",
                                                fg=self.renkler["ACCENT_GREEN"])
                    except Exception as e:
                        self.m_uyari_lbl.config(text=f"❌ Dosya kaydedilirken hata oluştu!",
                                                fg=self.renkler["ACCENT_RED"])
                return
            if format_tipi == "csv":
                dosya = filedialog.asksaveasfilename(defaultextension=".csv",
                                                     initialfile="Tasarruf Planı",
                                                     filetypes=[("Excel (CSV) Dosyası", "*.csv")],
                                                     title="Excel Olarak Kaydet")
                if dosya:
                    try:
                        with open(dosya, mode='w', newline='', encoding='utf-8-sig') as f:
                            writer = csv.writer(f, delimiter=';')
                            writer.writerow(
                                ["Dönem", "Vade Başlangıcı", "Vade Bitişi", "Net Faiz", "Nakit Akışı", "Nominal Bakiye",
                                 "Reel Bakiye"])
                            for satir in self.tree_m.get_children():
                                writer.writerow(self.tree_m.item(satir)['values'])
                        self.m_uyari_lbl.config(text=f"✅ Tasarruf planı Excel (CSV) olarak kaydedildi:\n{dosya}",
                                                fg=self.renkler["ACCENT_GREEN"])
                    except Exception:
                        self.m_uyari_lbl.config(text=f"❌ Dosya kaydedilirken hata oluştu!",
                                                fg=self.renkler["ACCENT_RED"])
            else:
                dosya = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                     initialfile="Tasarruf Planı",
                                                     filetypes=[("PDF Dosyası", "*.pdf")],
                                                     title="PDF Olarak Kaydet")
                if dosya:
                    try:
                        from reportlab.lib.pagesizes import A4, landscape
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors
                        from reportlab.pdfbase import pdfmetrics
                        from reportlab.pdfbase.ttfonts import TTFont

                        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
                        pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))

                        if self.karanlik_mod:
                            body_bg = colors.HexColor("#0D1117")
                            h2_renk = colors.HexColor("#58A6FF")
                            th_bg = colors.HexColor("#1F6FEB")
                            th_fg = colors.HexColor("#E6EDF3")
                            satir_cift = colors.HexColor("#1C2128")
                            satir_tek = colors.HexColor("#0D1117")
                            satir_fg = colors.HexColor("#E6EDF3")
                        else:
                            body_bg = colors.HexColor("#F0F4F8")
                            h2_renk = colors.HexColor("#0D47A1")
                            th_bg = colors.HexColor("#1A73E8")
                            th_fg = colors.white
                            satir_cift = colors.HexColor("#88A7B4")
                            satir_tek = colors.white
                            satir_fg = colors.HexColor("#0F172A")

                        doc = SimpleDocTemplate(dosya, pagesize=landscape(A4),
                                                leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
                        styles = getSampleStyleSheet()
                        story = []

                        baslik_style = styles["Title"]
                        baslik_style.textColor = h2_renk
                        baslik_style.fontName = "Arial-Bold"
                        story.append(Paragraph("Tasarruf Planı - Detaylı Ekstre", baslik_style))
                        story.append(Spacer(1, 12))

                        tablo_verisi = [
                            ["Dönem", "Vade Başlangıcı", "Vade Bitişi", "Net Faiz", "Nakit Akışı", "Nominal Bakiye",
                             "Reel Bakiye"]]
                        satir_verileri_pdf = [self.tree_m.item(s)['values'] for s in self.tree_m.get_children()]
                        for satir in satir_verileri_pdf:
                            tablo_verisi.append(list(satir))

                        tablo = Table(tablo_verisi, repeatRows=1)
                        style_cmds = [
                            ("BACKGROUND", (0, 0), (-1, 0), th_bg),
                            ("TEXTCOLOR", (0, 0), (-1, 0), th_fg),
                            ("FONTNAME", (0, 0), (-1, 0), "Arial-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 8),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [satir_cift, satir_tek]),
                            ("TEXTCOLOR", (0, 1), (-1, -1), satir_fg),
                            ("FONTNAME", (0, 1), (-1, -1), "Arial"),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("PADDING", (0, 0), (-1, -1), 6),
                        ]
                        tablo.setStyle(TableStyle(style_cmds))
                        story.append(tablo)
                        doc.build(story)
                        self.m_uyari_lbl.config(text=f"✅ Zebra tasarruf planı PDF olarak kaydedildi:\n{dosya}",
                                                fg=self.renkler["ACCENT_GREEN"])
                    except Exception as e:
                        self.m_uyari_lbl.config(text=f"❌ Dosya kaydedilirken hata oluştu: {str(e)}",
                                                fg=self.renkler["ACCENT_RED"])

        tk.Button(btn_frame, text="Excel (CSV)", bg=self.renkler["ACCENT_BLUE"], fg="#FFFFFF",
                  font=("Courier New", 10, "bold"), cursor="hand2", command=lambda: kaydet_baslat("csv")).pack(
            side="left", expand=True, fill="x", padx=5)
        tk.Button(btn_frame, text="Excel (XLSX) Zebra", bg="#217346", fg="#FFFFFF",
                  font=("Courier New", 10, "bold"), cursor="hand2", command=lambda: kaydet_baslat("xlsx")).pack(
            side="left", expand=True, fill="x", padx=5)
        tk.Button(btn_frame, text="PDF (Zebra)", bg=self.renkler["ACCENT_GREEN"], fg="#FFFFFF",
                  font=("Courier New", 10, "bold"), cursor="hand2", command=lambda: kaydet_baslat("html")).pack(
            side="right", expand=True, fill="x", padx=5)

    def _ozel_islem_ekle(self, donem_deger="", tutar_deger=""):
        r = self.renkler
        satir_frame = tk.Frame(self.ozel_islem_frame, bg=r["PANEL_BG"])
        satir_frame.pack(fill="x", pady=1)
        donem_var = tk.StringVar(value=str(donem_deger))
        tutar_var = tk.StringVar(value=str(tutar_deger))
        vcmd_int = (self.kok.register(lambda P: self._genel_dogrulama(P, "tamsayi")), '%P')
        vcmd_dec = (self.kok.register(lambda P: self._genel_dogrulama(P, "ondalik")), '%P')
        tk.Label(satir_frame, text="Dönem:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], font=("Courier New", 10)).pack(
            side="left")

        CustomEntry(satir_frame, is_dark=self.karanlik_mod, textvariable=donem_var, width=5,
                    bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                    insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 11), validate="key",
                    validatecommand=vcmd_int).pack(side="left", padx=(2, 6))

        tk.Label(satir_frame, text="Tutar (₺):", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], font=("Courier New", 10)).pack(
            side="left")

        CustomEntry(satir_frame, is_dark=self.karanlik_mod, textvariable=tutar_var, width=10,
                    bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                    insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]), font=("Courier New", 11), validate="key",
                    validatecommand=vcmd_dec).pack(side="left", padx=(2, 6))

        tk.Button(satir_frame, text="✕", bg=r["ACCENT_RED"], fg="#FFFFFF", relief="flat",
                  font=("Courier New", 9, "bold"), cursor="hand2",
                  command=lambda f=satir_frame, v=(donem_var, tutar_var): self._ozel_islem_sil(f, v)).pack(side="left")
        self.m_ozel_islemler_liste.append((donem_var, tutar_var))

    def _ozel_islem_sil(self, frame, var_tuple):
        if var_tuple in self.m_ozel_islemler_liste: self.m_ozel_islemler_liste.remove(var_tuple)
        frame.destroy()
        self.ozel_islem_frame.guncelle()

    def _borc_sekmesi(self, parent):
        r = self.renkler
        icerik = tk.Frame(parent, bg=r["DARK_BG"])
        icerik.pack(fill="both", expand=True, padx=8, pady=4)
        icerik.grid_columnconfigure(1, weight=1)
        icerik.grid_rowconfigure(0, weight=1)

        sol_panel = tk.Frame(icerik, bg=r["DARK_BG"])
        sol_panel.grid(row=0, column=0, sticky="ns", padx=(0, 16))
        sag_panel = tk.Frame(icerik, bg=r["DARK_BG"])
        sag_panel.grid(row=0, column=1, sticky="nsew")

        giris = tk.Frame(sol_panel, bg=r["PANEL_BG"], padx=16, pady=6)
        giris.pack(fill="x", pady=(0, 8))

        self.b_giris_acik = True

        def _b_toggle_giris():
            if self.b_giris_acik:
                giris.pack_forget()
                self.b_toggle_btn.config(text="▼  Parametreler", bg=r["ACCENT_DARK_BLUE"], fg="#FFFFFF")
                self.b_giris_acik = False
            else:
                giris.pack(fill="x", pady=(0, 8), before=self.b_toggle_btn_frame)
                self.b_toggle_btn.config(text="▲  Parametreleri Gizle", bg=r["PANEL_BG"], fg=r["ACCENT_BLUE"])
                self.b_giris_acik = True

        self.b_toggle_btn_frame = tk.Frame(sol_panel, bg=r["DARK_BG"])
        self.b_toggle_btn_frame.pack(fill="x", pady=(0, 4))
        self.b_toggle_btn = tk.Button(self.b_toggle_btn_frame, text="▲  Parametreleri Gizle",
                                      bg=r["PANEL_BG"], fg=r["ACCENT_BLUE"],
                                      font=("Courier New", 10, "bold"), relief="flat", cursor="hand2",
                                      command=_b_toggle_giris, anchor="w")
        self.b_toggle_btn.pack(fill="x")
        self._b_toggle_giris = _b_toggle_giris

        giris.grid_columnconfigure(0, minsize=260, weight=0)
        giris.grid_columnconfigure(1, weight=0)
        giris.grid_columnconfigure(2, weight=1)

        self.kredi_tip_frame = tk.Frame(giris, bg=r["ENTRY_BG"], highlightbackground=r["BORDER"], highlightthickness=1)
        self.kredi_tip_frame.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 12))
        self.kredi_sekmeleri = {}

        def _kredi_sekme_sec(tip_id, aktif_renk):
            eski_tip = self.b_kredi_tipi.get()

            if getattr(self, '_init_tamamlandi', False):
                if eski_tip in self.kredi_durumlari:
                    self.kredi_durumlari[eski_tip] = {
                        "teminat": self.b_teminat.get(), "baslangic": self.b_baslangic.get(),
                        "faiz": self.b_faiz.get(), "vergi": self.b_vergi.get(),
                        "odeme": self.b_odeme.get(), "vade_var": self.b_vade_var.get(),
                        "vade": self.b_vade.get(),
                        "ara_odemeler": [(d.get(), t.get()) for d, t in self.b_ara_odemeler_liste],
                        "hesaplandi": self.b_hesaplandi, "yeniden": self.b_yeniden_yapilandir_var.get()
                    }

            self.b_kredi_tipi.set(tip_id)
            for k, d in self.kredi_sekmeleri.items():
                d["btn"].config(bg=aktif_renk if k == tip_id else r["ENTRY_BG"],
                                fg=r["DARK_BG"] if k == tip_id else r["TEXT_MUTED"])

            if tip_id == "İhtiyaç Kredisi" or tip_id == "Özel Kredi":
                self.lbl_teminat_widget.grid_remove()
                self.frame_b_teminat.grid_remove()
                self.ikon_teminat.grid_remove()
            else:
                if tip_id == "Taşıt Kredisi":
                    self.lbl_teminat_text.set("Araç Değeri (₺):")
                    self.ikon_teminat.tooltip.text = "BDDK'nın LTV (Kredi/Değer) oranını hesaplayabilmesi için\naraç fatura değerini giriniz.\nMaksimum 10 basamaklı bir değer girin."
                elif tip_id == "Konut Kredisi":
                    self.lbl_teminat_text.set("Konut Değeri (₺):")
                    self.ikon_teminat.tooltip.text = "BDDK'nın LTV (Kredi/Değer) oranını hesaplayabilmesi için\nkonut ekspertiz değerini giriniz.\nMaksimum 10 basamaklı bir değer girin."
                self.lbl_teminat_widget.grid()
                self.frame_b_teminat.grid()
                self.ikon_teminat.grid()

            if tip_id in ["İhtiyaç Kredisi", "Taşıt Kredisi", "Konut Kredisi"]:
                self.entry_b_vergi.config(state="disabled")
            else:
                self.entry_b_vergi.config(state="normal")

            yeni_durum = self.kredi_durumlari[tip_id]
            self.b_teminat.set(yeni_durum["teminat"])
            self.b_baslangic.set(yeni_durum["baslangic"])
            self.b_faiz.set(yeni_durum["faiz"])
            self.b_vergi.set(yeni_durum["vergi"])
            self.b_odeme.set(yeni_durum["odeme"])
            self.b_vade_var.set(yeni_durum["vade_var"])
            self.b_yeniden_yapilandir_var.set(yeni_durum.get("yeniden", False))
            self._b_vade_toggle()
            self.b_vade.set(yeni_durum["vade"])

            for w in self.b_ara_odeme_frame.winfo_children(): w.destroy()
            self.b_ara_odemeler_liste.clear()
            for d_val, t_val in yeni_durum["ara_odemeler"]: self._b_ara_odeme_ekle(d_val, t_val)
            self._b_yapi_guncelle()

            self.b_hesaplandi = yeni_durum["hesaplandi"]

            if self.b_hesaplandi:
                self._borc_hesapla(sessiz=True)
            else:
                self._borc_temizle(sessiz=True)

            self._bddk_guncelle()
            self._dinamik_min_odeme()

        OZEL_RENK = "#C9A000" if not self.karanlik_mod else "#D4B000"
        for tip_id, text, renk in [("İhtiyaç Kredisi", "💸 İhtiyaç", r["ACCENT_GREEN"]),
                                   ("Taşıt Kredisi", "🚗 Taşıt", r["ACCENT_ORANGE"]),
                                   ("Konut Kredisi", "🏠 Konut", r["ACCENT_BLUE"]),
                                   ("Özel Kredi", "🛠️ Özel", OZEL_RENK)]:
            btn = tk.Button(self.kredi_tip_frame, text=text, font=("Courier New", 11, "bold"), bg=r["ENTRY_BG"],
                            fg=r["TEXT_MUTED"], relief="flat", cursor="hand2", bd=0,
                            command=lambda t=tip_id, c=renk: _kredi_sekme_sec(t, c))
            btn.pack(side="left", fill="x", expand=True, ipady=6)
            self.kredi_sekmeleri[tip_id] = {"btn": btn, "renk": renk}

        tk.Label(giris, text="Başlangıç Tarihi:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                 font=("Courier New", 11), width=28, anchor="w").grid(row=1, column=0, sticky="w", pady=2, padx=(0, 8))
        self._tarih_entry_olustur(giris, self.b_tarih, 1)

        self.lbl_teminat_text = tk.StringVar(value="Teminat Değeri (₺):")
        self.lbl_teminat_widget = tk.Label(giris, textvariable=self.lbl_teminat_text, bg=r["PANEL_BG"],
                                           fg=r["TEXT_PRIMARY"], font=("Courier New", 11), width=28, anchor="w")
        self.lbl_teminat_widget.grid(row=2, column=0, sticky="w", pady=2, padx=(0, 8))

        vcmd_gorsel = (self.kok.register(lambda P: self._genel_dogrulama(P, "para_gorsel")), '%P')

        self.frame_b_teminat = tk.Frame(giris, bg=r["PANEL_BG"])
        self.frame_b_teminat.grid(row=2, column=1, sticky="w", pady=2)
        self.entry_b_teminat = CustomEntry(self.frame_b_teminat, is_dark=self.karanlik_mod, textvariable=self.b_teminat,
                                           width=15,
                                           bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]),
                                           fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                                           insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                                           font=("Courier New", 12), validate="key", validatecommand=vcmd_gorsel)
        self.entry_b_teminat.pack(side="left")
        self.entry_b_teminat.is_currency_formatted = True
        self.entry_b_teminat.bind('<KeyRelease>', self._para_formatla_event, add='+')
        self.ikon_teminat = self._ikon_olustur(giris,
                                               "BDDK'nın LTV (Kredi/Değer) oranını hesaplayabilmesi için\naraç fatura değeri veya konut ekspertiz değerini girin.")
        self.ikon_teminat.grid(row=2, column=2, sticky="w", padx=(4, 0))

        entry_b_bas = self._etiket_giris(giris, "Talep Edilen Kredi (₺):", "b_baslangic", 3, "para_gorsel",
                                         "Çekmek istediğiniz borç/kredi tutarını giriniz.\nMaksimum 10 basamaklı bir değer girin.")
        entry_b_bas.is_currency_formatted = True
        entry_b_bas.bind('<KeyRelease>', self._para_formatla_event, add='+')

        self.lbl_bddk_bilgi = tk.Label(giris, text="", bg=r["PANEL_BG"], fg=r["ACCENT_BLUE"],
                                       font=("Courier New", 10, "bold"), justify="left", wraplength=430)
        self.lbl_bddk_bilgi.grid(row=4, column=0, columnspan=3, sticky="w", pady=(0, 6), padx=(0, 8))

        self.entry_b_faiz = self._etiket_giris(giris, "Kredi Faiz Oranı (Yıllık %):", "b_faiz", 5, "yuzde",
                                               "Kredinin yıllık brüt faiz oranını giriniz.\n%0.0 - %100.0 arasında bir değer girin.")
        self.entry_b_vergi = self._etiket_giris(giris, "Vergi (KKDF+BSMV %):", "b_vergi", 6, "yuzde",
                                                "Faize eklenen yasal vergi oranını giriniz.\n%0.0 - %100.0 arasında bir değer girin.")
        self.entry_b_odeme = self._etiket_giris(giris, "Aylık Ödeme (₺):", "b_odeme", 7, "para_gorsel",
                                                "Aylık sabit taksit tutarını giriniz.\nAylık ödeme en fazla toplam borca eşit olabilir.")
        self.entry_b_odeme.is_currency_formatted = True
        self.entry_b_odeme.bind('<KeyRelease>', self._para_formatla_event, add='+')

        f_ara_btn = tk.Frame(giris, bg=r["PANEL_BG"])
        f_ara_btn.grid(row=8, column=0, sticky="w", pady=(6, 2))
        tk.Label(f_ara_btn, text="Ara Ödemeler:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"],
                 font=("Courier New", 11)).pack(side="left")
        self._ikon_olustur(f_ara_btn,
                           "Aylık ödemeler dışında yapacağınız\nara ödemeleri varsa belirtiniz.\n(Para yatırma işlemi için örn: 3. ay 800 şeklinde yazınız.)").pack(
            side="left", padx=(4, 0))

        self.btn_ara_ekle = tk.Button(giris, text="＋ Ekle", bg=r["ACCENT_BLUE"], fg=r["DARK_BG"], relief="flat",
                                      font=("Courier New", 9, "bold"), cursor="hand2", command=self._b_ara_odeme_ekle)
        self.btn_ara_ekle.grid(row=8, column=1, sticky="e")

        self.b_ara_odeme_container, self.b_ara_odeme_frame = self._scrollable_frame_olustur(giris)
        self.b_ara_odeme_container.grid(row=9, column=0, columnspan=3, sticky="ew", pady=(0, 4))

        self.b_vade_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        self.b_vade_frame.grid(row=10, column=0, sticky="w", pady=4)
        tk.Checkbutton(self.b_vade_frame, text="Vade Hedefi (Ay):", variable=self.b_vade_var,
                       command=self._b_vade_toggle, bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], selectcolor=r["ENTRY_BG"],
                       activebackground=r["PANEL_BG"], activeforeground=r["TEXT_PRIMARY"],
                       font=("Courier New", 11)).pack(side="left", padx=(0, 8))

        vcmd_int = (self.kok.register(lambda P: self._genel_dogrulama(P, "tamsayi")), '%P')

        self.b_vade_entry_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        self.b_vade_entry = CustomEntry(self.b_vade_entry_frame, is_dark=self.karanlik_mod, textvariable=self.b_vade,
                                        width=8,
                                        bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]),
                                        fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                                        insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                                        font=("Courier New", 12), validate="key", validatecommand=vcmd_int)
        self.b_vade_entry.pack(side="left")
        if self.b_vade_var.get(): self.b_vade_entry_frame.grid(row=10, column=1, sticky="w", pady=4)
        self._ikon_olustur(giris,
                           "Borcunuzu hesaplanan dönemden farklı bir dönemde bitirmek\nistiyorsanız belirtiniz.\nAylık ödemeniz bu döneme göre hesaplanır.").grid(
            row=10, column=2, sticky="w", padx=(6, 0))

        # --- YENİ EKLENEN KISIM: Yeniden Yapılandırma ---
        self.b_yapi_frame = tk.Frame(giris, bg=r["PANEL_BG"])
        tk.Checkbutton(self.b_yapi_frame, text="Ara ödemede taksiti düşür (Vade sabit kalır)",
                       variable=self.b_yeniden_yapilandir_var,
                       bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], selectcolor=r["ENTRY_BG"],
                       activebackground=r["PANEL_BG"], activeforeground=r["TEXT_PRIMARY"],
                       font=("Courier New", 10)).pack(side="left", padx=(0, 8))
        self._ikon_olustur(self.b_yapi_frame,
                           "İşaretli Değilse: Ara ödeme sonrası taksit sabit kalır ama borç ERKEN BİTER.\nİşaretliyse: Vade sabit kalır, ancak kalan aylık taksitleriniz DÜŞER.").pack(
            side="left", padx=(4, 0))
        # grid işlemi _b_yapi_guncelle içinden yapılıyor

        self.lbl_dinamik_min = tk.Label(giris, text="", bg=r["PANEL_BG"], fg=r["ACCENT_TAHAKKUK"],
                                        font=("Courier New", 10, "bold"), justify="left", wraplength=430)
        self.lbl_dinamik_min.grid(row=12, column=0, columnspan=3, sticky="w", pady=(2, 0), padx=(0, 8))
        self.lbl_vade_min_odeme = tk.Label(giris, text="", bg=r["PANEL_BG"], fg=r["ACCENT_BLUE"],
                                           font=("Courier New", 10, "bold"), justify="left", wraplength=430)
        self.lbl_vade_min_odeme.grid(row=13, column=0, columnspan=3, sticky="w", pady=(0, 8), padx=(0, 8))

        self.btn_frame_b = tk.Frame(giris, bg=r["PANEL_BG"])
        self.btn_frame_b.grid(row=14, column=0, columnspan=3, sticky="ew", pady=(8, 0))

        self.btn_frame_b.grid_columnconfigure(0, weight=4)
        self.btn_frame_b.grid_columnconfigure(1, weight=1)
        self.btn_frame_b.grid_columnconfigure(2, weight=0)

        tk.Button(self.btn_frame_b, text="  Hesapla  ", bg=r["ACCENT_BLUE"], fg="#FFFFFF",
                  font=("Courier New", 11, "bold"),
                  relief="flat", cursor="hand2", command=self._borc_hesapla).grid(row=0, column=0, sticky="ew")
        tk.Button(self.btn_frame_b, text="Temizle", bg=r["ACCENT_RED"], fg="#FFFFFF", font=("Courier New", 11, "bold"),
                  relief="flat", cursor="hand2", command=self._borc_temizle).grid(row=0, column=1, sticky="ew",
                                                                                  padx=(4, 0))

        self.btn_indir_csv = tk.Button(self.btn_frame_b, text="İndir", bg=r["ACCENT_GREEN"], fg="#FFFFFF",
                                       font=("Courier New", 11, "bold"),
                                       relief="flat", cursor="hand2", command=self._indirme_popup_ac)
        self.btn_indir_csv.grid(row=0, column=2, sticky="ew", padx=(4, 0))
        self.btn_indir_csv.grid_remove()

        self.b_sonuc_container, self.b_sonuc_frame = self._standart_scrollable_frame(sol_panel, bg_color=r["PANEL_BG"],
                                                                                     padx=16, pady=8)
        self.b_sonuc_container.pack(fill="both", expand=True, pady=(0, 8))

        self.b_uyari_lbl = tk.Label(self.b_sonuc_frame, text="", bg=r["PANEL_BG"], fg=r["ACCENT_ORANGE"],
                                    font=("Courier New", 11, "bold"), justify="left", anchor="nw")
        self.b_uyari_lbl.pack(anchor="w", pady=(0, 8), fill="x")
        self.b_uyari_lbl.bind("<Configure>", lambda e: e.widget.config(wraplength=max(50, e.width - 5)))

        self.b_sonuc_grid_frame = tk.Frame(self.b_sonuc_frame, bg=r["PANEL_BG"])
        self.b_sonuc_grid_frame.pack(anchor="w", fill="x")

        sag_panel.grid_rowconfigure(0, weight=1)
        sag_panel.grid_rowconfigure(1, weight=2)
        sag_panel.grid_columnconfigure(0, weight=1)
        self.b_canvas = tk.Canvas(sag_panel, bg=r["CHART_BG"], highlightthickness=0)
        self.b_canvas.grid(row=0, column=0, sticky="nsew", pady=(0, 10))

        tablo_frame = tk.Frame(sag_panel, bg=r["ENTRY_BG"])
        tablo_frame.grid(row=1, column=0, sticky="nsew")
        tablo_frame.grid_rowconfigure(0, weight=1)
        tablo_frame.grid_rowconfigure(1, weight=0)
        tablo_frame.grid_columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(tablo_frame,
                                 columns=("taksit", "tarih", "kalan", "faiz", "anapara", "ara_odeme", "tutar"),
                                 show="headings")
        for c, n, w in [("taksit", "No", 50), ("tarih", "Tarih", 100), ("kalan", "Kalan Anapara", 120),
                        ("faiz", "Faiz Yükü", 100), ("anapara", "Anapara Öd.", 100), ("ara_odeme", "Ara Ödeme", 100),
                        ("tutar", "Toplam Taksit", 120)]:
            self.tree.heading(c, text=n)
            self.tree.column(c, width=w, anchor="center" if c in ("taksit", "tarih") else "e")
        self.tree.tag_configure("evenrow", background=r["ROW_EVEN"])
        self.tree.tag_configure("oddrow", background=r["ROW_ODD"])

        scroll = ttk.Scrollbar(tablo_frame, orient="vertical", command=self.tree.yview)
        scroll_x = ttk.Scrollbar(tablo_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scroll.set, xscrollcommand=scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<<TreeviewSelect>>", lambda e: self._tablo_secim_grafik(self.tree, self.b_canvas, "b"))

        mevcut_tip = self.b_kredi_tipi.get()
        renk = self.kredi_sekmeleri.get(mevcut_tip, {}).get("renk", r["ACCENT_GREEN"])
        _kredi_sekme_sec(mevcut_tip, renk)

    def _indirme_popup_ac(self):
        if not self.b_hesaplandi:
            self.b_uyari_lbl.config(text="❌ Lütfen önce borç hesaplaması yapınız!", fg=self.renkler["ACCENT_RED"])
            return

        top = tk.Toplevel(self.kok)
        top.title("İndirme Formatı")
        top.geometry("480x150")
        top.configure(bg=self.renkler["PANEL_BG"])
        top.transient(self.kok)
        top.grab_set()

        top.update_idletasks()
        x = self.kok.winfo_rootx() + (self.kok.winfo_width() - top.winfo_width()) // 2
        y = self.kok.winfo_rooty() + (self.kok.winfo_height() - top.winfo_height()) // 2
        top.geometry(f"+{x}+{y}")

        tk.Label(top, text="Hangi formatta indirmek istersiniz?", bg=self.renkler["PANEL_BG"],
                 fg=self.renkler["TEXT_PRIMARY"], font=("Courier New", 11, "bold")).pack(pady=25)

        btn_frame = tk.Frame(top, bg=self.renkler["PANEL_BG"])
        btn_frame.pack(fill="x", padx=20)

        def kaydet_baslat(format_tipi):
            top.destroy()
            tip = self.b_kredi_tipi.get()
            dosya_adi = {
                "İhtiyaç Kredisi": "İhtiyaç Kredisi Ödeme Planı",
                "Taşıt Kredisi": "Araç Kredisi Ödeme Planı",
                "Konut Kredisi": "Özel Kredi Ödeme Planı"
            }.get(tip, "Ödeme Planı")

            if format_tipi == "xlsx":
                dosya = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                     initialfile=dosya_adi,
                                                     filetypes=[("Excel Dosyası", "*.xlsx")],
                                                     title="Excel Olarak Kaydet")
                if dosya:
                    try:
                        basliklar = ["Taksit No", "Tarih", "Kalan Anapara", "Faiz Yükü",
                                     "Anapara Ödemesi", "Ara Ödeme", "Toplam Taksit"]
                        satir_verileri = [self.tree.item(s)['values'] for s in self.tree.get_children()]
                        self._xlsx_zebra_kaydet(dosya, basliklar, satir_verileri,
                                                sayfa_adi="Ödeme Planı",
                                                tablo_basligi=f"{dosya_adi} - Detaylı Ekstre")
                        self.b_uyari_lbl.config(text=f"✅ Zebra ödeme planı Excel olarak kaydedildi:\n{dosya}",
                                                fg=self.renkler["ACCENT_GREEN"])
                    except Exception:
                        self.b_uyari_lbl.config(text=f"❌ Dosya kaydedilirken hata oluştu!",
                                                fg=self.renkler["ACCENT_RED"])
                return

            if format_tipi == "csv":
                dosya = filedialog.asksaveasfilename(defaultextension=".csv",
                                                     initialfile=dosya_adi,
                                                     filetypes=[("Excel (CSV) Dosyası", "*.csv")],
                                                     title="Excel Olarak Kaydet")
                if dosya:
                    try:
                        with open(dosya, mode='w', newline='', encoding='utf-8-sig') as f:
                            writer = csv.writer(f, delimiter=';')
                            writer.writerow(
                                ["Taksit No", "Tarih", "Kalan Anapara", "Faiz Yükü", "Anapara Ödemesi", "Ara Ödeme",
                                 "Toplam Taksit"])
                            for satir in self.tree.get_children():
                                writer.writerow(self.tree.item(satir)['values'])
                        self.b_uyari_lbl.config(text=f"✅ Ödeme planı Excel (CSV) olarak kaydedildi:\n{dosya}",
                                                fg=self.renkler["ACCENT_GREEN"])
                    except Exception:
                        self.b_uyari_lbl.config(text=f"❌ Dosya kaydedilirken hata oluştu!",
                                                fg=self.renkler["ACCENT_RED"])
            else:
                dosya = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                     initialfile=dosya_adi,
                                                     filetypes=[("PDF Dosyası", "*.pdf")],
                                                     title="PDF Olarak Kaydet")
                if dosya:
                    try:
                        from reportlab.lib.pagesizes import A4, landscape
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                        from reportlab.lib.styles import getSampleStyleSheet
                        from reportlab.lib import colors
                        from reportlab.pdfbase import pdfmetrics
                        from reportlab.pdfbase.ttfonts import TTFont

                        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
                        pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))

                        if self.karanlik_mod:
                            body_bg = colors.HexColor("#0D1117")
                            h2_renk = colors.HexColor("#58A6FF")
                            th_bg = colors.HexColor("#1F6FEB")
                            th_fg = colors.HexColor("#E6EDF3")
                            satir_cift = colors.HexColor("#1C2128")
                            satir_tek = colors.HexColor("#0D1117")
                            satir_fg = colors.HexColor("#E6EDF3")
                        else:
                            body_bg = colors.HexColor("#F0F4F8")
                            h2_renk = colors.HexColor("#0D47A1")
                            th_bg = colors.HexColor("#1A73E8")
                            th_fg = colors.white
                            satir_cift = colors.HexColor("#88A7B4")
                            satir_tek = colors.white
                            satir_fg = colors.HexColor("#0F172A")

                        doc = SimpleDocTemplate(dosya, pagesize=landscape(A4),
                                                leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
                        styles = getSampleStyleSheet()
                        story = []

                        baslik_style = styles["Title"]
                        baslik_style.textColor = h2_renk
                        baslik_style.fontName = "Arial-Bold"
                        story.append(Paragraph(f"{dosya_adi} - Detaylı Ekstre", baslik_style))
                        story.append(Spacer(1, 12))

                        tablo_verisi = [["Taksit No", "Tarih", "Kalan Anapara", "Faiz Yükü",
                                         "Anapara Ödemesi", "Ara Ödeme", "Toplam Taksit"]]
                        satir_verileri_pdf = [self.tree.item(s)['values'] for s in self.tree.get_children()]
                        for satir in satir_verileri_pdf:
                            tablo_verisi.append(list(satir))

                        tablo = Table(tablo_verisi, repeatRows=1)
                        style_cmds = [
                            ("BACKGROUND", (0, 0), (-1, 0), th_bg),
                            ("TEXTCOLOR", (0, 0), (-1, 0), th_fg),
                            ("FONTNAME", (0, 0), (-1, 0), "Arial-Bold"),
                            ("FONTSIZE", (0, 0), (-1, -1), 8),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [satir_cift, satir_tek]),
                            ("TEXTCOLOR", (0, 1), (-1, -1), satir_fg),
                            ("FONTNAME", (0, 1), (-1, -1), "Arial"),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                            ("PADDING", (0, 0), (-1, -1), 6),
                        ]
                        tablo.setStyle(TableStyle(style_cmds))
                        story.append(tablo)
                        doc.build(story)
                        self.b_uyari_lbl.config(text=f"✅ Zebra ödeme planı PDF olarak kaydedildi:\n{dosya}",
                                                fg=self.renkler["ACCENT_GREEN"])
                    except Exception as e:
                        self.b_uyari_lbl.config(text=f"❌ Dosya kaydedilirken hata oluştu: {str(e)}",
                                                fg=self.renkler["ACCENT_RED"])

        tk.Button(btn_frame, text="Excel (CSV)", bg=self.renkler["ACCENT_BLUE"], fg="#FFFFFF",
                  font=("Courier New", 10, "bold"), cursor="hand2", command=lambda: kaydet_baslat("csv")).pack(
            side="left", expand=True, fill="x", padx=5)
        tk.Button(btn_frame, text="Excel (XLSX) Zebra", bg="#217346", fg="#FFFFFF",
                  font=("Courier New", 10, "bold"), cursor="hand2", command=lambda: kaydet_baslat("xlsx")).pack(
            side="left", expand=True, fill="x", padx=5)
        tk.Button(btn_frame, text="PDF (Zebra)", bg=self.renkler["ACCENT_GREEN"], fg="#FFFFFF",
                  font=("Courier New", 10, "bold"), cursor="hand2", command=lambda: kaydet_baslat("html")).pack(
            side="right", expand=True, fill="x", padx=5)

    def _xlsx_zebra_kaydet(self, dosya, basliklar, satir_verileri, sayfa_adi="Ekstre", tablo_basligi=""):
        TH_BG = "1A73E8"
        TH_FG = "FFFFFF"
        ROW_CIFT = "FFFFFF"
        ROW_TEK = "88A7B4"
        H2_FG = "0D47A1"
        BORDER_C = "CBD5E1"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sayfa_adi

        thin = Side(style="thin", color=BORDER_C)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        if tablo_basligi:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(basliklar))
            hucre = ws.cell(row=1, column=1, value=tablo_basligi)
            hucre.font = Font(name="Arial", bold=True, size=13, color=H2_FG)
            hucre.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 24
            veri_baslangic = 3
        else:
            veri_baslangic = 2

        baslik_satir = veri_baslangic - 1
        for col_idx, baslik in enumerate(basliklar, start=1):
            h = ws.cell(row=baslik_satir, column=col_idx, value=baslik)
            h.fill = PatternFill("solid", fgColor=TH_BG)
            h.font = Font(name="Arial", bold=True, color=TH_FG, size=11)
            h.alignment = Alignment(horizontal="center", vertical="center")
            h.border = border
        ws.row_dimensions[baslik_satir].height = 22

        for i, satir in enumerate(satir_verileri):
            row_idx = veri_baslangic + i
            bg = ROW_CIFT if i % 2 == 0 else ROW_TEK
            fill = PatternFill("solid", fgColor=bg)
            for col_idx, deger in enumerate(satir, start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=str(deger))
                c.fill = fill
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(horizontal="center" if col_idx <= 2 else "right", vertical="center")
                c.border = border
            ws.row_dimensions[row_idx].height = 20

        for col_idx, baslik in enumerate(basliklar, start=1):
            maks = max(
                len(baslik),
                *[len(str(satir[col_idx - 1])) for satir in satir_verileri] if satir_verileri else [0]
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(maks + 4, 30)

        wb.save(dosya)

    def _b_ara_odeme_sil(self, frame, var_tuple):
        if var_tuple in self.b_ara_odemeler_liste: self.b_ara_odemeler_liste.remove(var_tuple)
        frame.destroy()
        self.b_ara_odeme_frame.guncelle()
        self._b_yapi_guncelle()

    def _b_ara_odeme_ekle(self, d_val="", t_val=""):
        r = self.renkler
        f = tk.Frame(self.b_ara_odeme_frame, bg=r["PANEL_BG"])
        f.pack(fill="x", pady=1)
        d_var = tk.StringVar(value=str(d_val))
        t_val_clean = str(t_val).lstrip('-') if str(t_val).startswith('-') else str(t_val)
        t_var = tk.StringVar(value=t_val_clean)

        def _eksi_engelle(*args):
            val = t_var.get()
            if val.startswith('-'):
                t_var.set(val.lstrip('-'))

        t_var.trace_add("write", _eksi_engelle)

        tk.Label(f, text="Ay:", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], font=("Courier New", 10)).pack(side="left")

        CustomEntry(f, is_dark=self.karanlik_mod, textvariable=d_var, width=5,
                    bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                    insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"])).pack(side="left", padx=(2, 6))

        tk.Label(f, text="Tutar (₺):", bg=r["PANEL_BG"], fg=r["TEXT_PRIMARY"], font=("Courier New", 10)).pack(
            side="left")

        CustomEntry(f, is_dark=self.karanlik_mod, textvariable=t_var, width=10,
                    bg=r.get("TEXTBOX_BG", r["ENTRY_BG"]), fg=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"]),
                    insertbackground=r.get("TEXTBOX_FG", r["TEXT_PRIMARY"])).pack(side="left", padx=(2, 6))

        tk.Button(f, text="✕", bg=r["ACCENT_RED"], fg="#FFFFFF", relief="flat", font=("Courier New", 9, "bold"),
                  cursor="hand2",
                  command=lambda frm=f, v=(d_var, t_var): self._b_ara_odeme_sil(frm, v)).pack(side="left")
        self.b_ara_odemeler_liste.append((d_var, t_var))
        self._b_yapi_guncelle()

    def _mevduat_temizle(self):
        for var in [self.m_baslangic, self.m_faiz, self.m_stopaj, self.m_enflasyon, self.m_vade_gun, self.m_yatirim,
                    self.m_sure]: var.set("")
        for w in self.erken_cekim_icerik.winfo_children(): w.destroy()
        self.m_erken_cekim_liste.clear()
        self.erken_cekim_icerik.guncelle()
        self.m_duzenli_islem_var.set(False);
        self._duzenli_islem_toggle()
        self.m_enflasyon_aktif.set(False);
        self._enflasyon_toggle()
        for w in self.ozel_islem_frame.winfo_children(): w.destroy()
        self.m_ozel_islemler_liste.clear()
        self.ozel_islem_frame.guncelle()

        if not self.m_giris_acik:
            self._m_toggle_giris()

        for w in self.m_sonuc_grid_frame.winfo_children(): w.destroy()
        self.m_uyari_lbl.config(text="")

        self.tree_m.delete(*self.tree_m.get_children())
        if hasattr(self, 'm_tablo_not_lbl'):
            self.m_tablo_not_lbl.config(text="")
        self._bosh_grafik_ciz(self.m_canvas, self.renkler, "Tasarruf Eğrisi")
        self.m_hesaplandi = False

        if hasattr(self, 'btn_indir_csv_m'):
            self.btn_indir_csv_m.grid_remove()
            self.btn_frame_m.grid_columnconfigure(0, weight=4)
            self.btn_frame_m.grid_columnconfigure(1, weight=1)
            self.btn_frame_m.grid_columnconfigure(2, weight=0)

    def _borc_temizle(self, sessiz=False):
        tip_id = self.b_kredi_tipi.get()
        if not sessiz:
            for var in [self.b_teminat, self.b_baslangic, self.b_faiz, self.b_odeme, self.b_vade]: var.set("")

            if tip_id == "Özel Kredi":
                self.b_vergi.set("")
            elif tip_id in ["İhtiyaç Kredisi", "Taşıt Kredisi"]:
                self.b_vergi.set("30")
            elif tip_id == "Konut Kredisi":
                self.b_vergi.set("0")

            self.b_vade_var.set(False);
            self._b_vade_toggle()
            for w in self.b_ara_odeme_frame.winfo_children(): w.destroy()
            self.b_ara_odemeler_liste.clear()
            self.b_ara_odeme_frame.guncelle()
            self._b_yapi_guncelle()

            if not self.b_giris_acik:
                self._b_toggle_giris()

        for w in self.b_sonuc_grid_frame.winfo_children(): w.destroy()
        self.b_uyari_lbl.config(text="")

        self.tree.delete(*self.tree.get_children())
        self._bosh_grafik_ciz(self.b_canvas, self.renkler, "Borç Eğrisi")
        self.b_hesaplandi = False

        if hasattr(self, 'btn_indir_csv'):
            self.btn_indir_csv.grid_remove()
            self.btn_frame_b.grid_columnconfigure(0, weight=4)
            self.btn_frame_b.grid_columnconfigure(1, weight=1)
            self.btn_frame_b.grid_columnconfigure(2, weight=0)

        if getattr(self, '_init_tamamlandi', False) and tip_id in self.kredi_durumlari and not sessiz:
            self.kredi_durumlari[tip_id] = {"teminat": "", "baslangic": "", "faiz": "", "vergi": self.b_vergi.get(),
                                            "odeme": "", "vade_var": False, "vade": "", "ara_odemeler": [],
                                            "hesaplandi": False, "yeniden": False}
        if not sessiz:
            self._bddk_guncelle()
            self._dinamik_min_odeme()

    def _tablo_secim_grafik(self, tree, canvas, mod):
        secim = tree.selection()
        if not secim: return

        degerler = tree.item(secim[0])['values']
        if not degerler: return

        donem_metni = str(degerler[0])
        try:
            hedef_ay = int(donem_metni.split('-')[0])
        except ValueError:
            return

        canvas.delete("tablo_vurgu")
        canvas.itemconfig("tooltip_text", state="hidden")
        canvas.itemconfig("tooltip_bg", state="hidden")
        if not hasattr(canvas, "noktalar") or not canvas.noktalar: return

        hedef_nokta = next((p for p in canvas.noktalar if p[3] == hedef_ay), None)
        if hedef_nokta:
            px, py, val, _ = hedef_nokta
            H = canvas.winfo_height() if canvas.winfo_height() > 50 else 300
            canvas.create_line(px, 50, px, H - 60, fill=self.renkler["ACCENT_GOLD"], width=2, dash=(4, 2),
                               tags="tablo_vurgu")
            canvas.create_oval(px - 6, py - 6, px + 6, py + 6, fill=self.renkler["ACCENT_GOLD"],
                               outline=self.renkler["PANEL_BG"], width=2, tags="tablo_vurgu")
            canvas.itemconfig("tooltip_text", text=f"Dönem: {hedef_ay}\nDeğer: ₺{val:,.2f}", state="normal")
            W = canvas.winfo_width() if canvas.winfo_width() > 50 else 750
            canvas.coords("tooltip_text", px + (-60 if px > W - 150 else 30), py - 20)
            bbox = canvas.bbox("tooltip_text")
            if bbox: canvas.coords("tooltip_bg", bbox[0] - 6, bbox[1] - 6, bbox[2] + 6, bbox[3] + 6); canvas.itemconfig(
                "tooltip_bg", state="normal")
            canvas.tag_raise("tablo_vurgu")
            canvas.tag_raise("tooltip_bg")
            canvas.tag_raise("tooltip_text")

    def _hata_ve_temizle(self, var, baslik, mesaj, tab_lbl=None):
        if tab_lbl:
            self._dinamik_yazdir(tab_lbl, f"❌ {baslik}: {mesaj}", base_size=11, is_bold=True)
            tab_lbl.config(fg=self.renkler["ACCENT_RED"])
        if var: var.set("")

    def _dogrula(self, deger_str, min_val, max_val, ad, lbl_uyari=None, sessiz=False):
        try:
            if deger_str in ["", ".", "-"]: raise ValueError
            val = float(deger_str)
            if not (min_val <= val <= max_val): raise ValueError
            return val
        except ValueError:
            if not sessiz and lbl_uyari: self._hata_ve_temizle(None, "Geçersiz Veri",
                                                               f"'{ad}' alanına {min_val} ile {max_val} arası eksiksiz sayı girin.",
                                                               lbl_uyari)
            return None

    def _dogrula_int(self, deger_str, min_val, max_val, ad, lbl_uyari=None, sessiz=False):
        try:
            if deger_str == "": raise ValueError
            val = int(deger_str)
            if not (min_val <= val <= max_val): raise ValueError
            return val
        except ValueError:
            if not sessiz and lbl_uyari: self._hata_ve_temizle(None, "Geçersiz Veri",
                                                               f"'{ad}' alanına {min_val}-{max_val} arası geçerli TAM SAYI girin.",
                                                               lbl_uyari)
            return None

    def _tarih_dogrula(self, tarih_str, lbl_uyari=None, sessiz=False):
        try:
            return datetime.datetime.strptime(tarih_str, "%d.%m.%Y").date()
        except ValueError:
            if not sessiz and lbl_uyari: self._hata_ve_temizle(None, "Tarih Hatası",
                                                               "Tarihi GG.AA.YYYY formatında eksiksiz girin.",
                                                               lbl_uyari)
            return None

    def _parse_ozel_islemler(self, lbl_uyari, sessiz):
        islemler = {}
        for donem_var, tutar_var in self.m_ozel_islemler_liste:
            d_str = donem_var.get().strip()
            t_str = tutar_var.get().strip()
            if not d_str: continue
            if d_str and not t_str: t_str = "0"
            try:
                d = int(d_str)
                t = float(t_str)
                if d < 1: raise ValueError
                islemler[d] = Decimal(str(t))
            except (ValueError, Exception):
                if not sessiz and lbl_uyari: self._hata_ve_temizle(None, "Özel İşlem Hatası",
                                                                   "İşlem satırlarını doğru doldurun (Dönem: tam sayı).",
                                                                   lbl_uyari)
                return None
        return islemler

    def _mevduat_hesapla(self, sessiz=False):
        r = self.renkler
        self.m_uyari_lbl.config(text="")
        lbl = self.m_uyari_lbl

        for w in self.m_sonuc_grid_frame.winfo_children(): w.destroy()
        self.tree_m.delete(*self.tree_m.get_children())
        if hasattr(self, 'm_tablo_not_lbl'): self.m_tablo_not_lbl.config(text="")

        faiz_str = self.m_faiz.get()
        if faiz_str in ["", ".", "-"]: faiz_str = "0"
        faiz = self._dogrula(faiz_str, 0, 100, "Brüt Faiz", lbl, sessiz)
        if faiz is None: return False

        vade_str = self.m_vade_gun.get()
        if vade_str == "" or vade_str == "0":
            if not sessiz: self._hata_ve_temizle(self.m_vade_gun, "Eksik Veri", "'Vade Günü' 1'den küçük olamaz.", lbl)
            return False
        vade = self._dogrula_int(vade_str, 1, 999, "Vade Günü", lbl, sessiz)
        if vade is None: return False

        stopaj = float(stopaj_hesapla(vade))
        self.m_stopaj.set(str(stopaj).rstrip('0').rstrip('.') if '.' in str(stopaj) else str(stopaj))

        sure_str = self.m_sure.get()
        if sure_str == "" or sure_str == "0":
            if not sessiz: self._hata_ve_temizle(self.m_sure, "Eksik Veri", "'Tekrar Edecek Dönem' 1'den küçük olamaz.",
                                                 lbl)
            return False
        N = self._dogrula_int(sure_str, 1, 120, "Dönem", lbl, sessiz)
        if N is None: return False

        tarih = self._tarih_dogrula(self.m_tarih.get(), lbl, sessiz)
        if tarih is None: return False

        baslangic_str = self.m_baslangic.get().replace('.', '').replace(',', '.')
        if baslangic_str in ["", ".", "-", "-."]: baslangic_str = "0"
        try:
            A0 = float(baslangic_str)
        except:
            if not sessiz: self._hata_ve_temizle(self.m_baslangic, "Hata", "Geçersiz Başlangıç Bakiye formati.", lbl)
            return False

        if A0 == 0 and faiz == 0:
            if not sessiz: self._hata_ve_temizle(self.m_faiz, "Hata", "Başlangıç bakiyesi 0 olduğundan faiz 0 olamaz.",
                                                 lbl)
            return False

        if self.m_duzenli_islem_var.get():
            yatirim_str = self.m_yatirim.get().replace('.', '').replace(',', '.')
            if yatirim_str in ["", ".", "-", "-."]: yatirim_str = "0"
            try:
                P = float(yatirim_str)
            except:
                if not sessiz: self._hata_ve_temizle(self.m_yatirim, "Hata", "Geçersiz Düzenli İşlem formati.", lbl)
                return False
        else:
            P = 0.0

        enf_str = self.m_enflasyon.get()
        if enf_str in ["", ".", "-"]: enf_str = "0"
        try:
            enf = float(enf_str) if self.m_enflasyon_aktif.get() else 0.0
        except ValueError:
            if not sessiz: self._hata_ve_temizle(self.m_enflasyon, "Hata", "Enflasyon formatı hatalı.", lbl)
            return False

        ozel_islem_dict = self._parse_ozel_islemler(lbl, sessiz)
        if ozel_islem_dict is None: return False

        erken_cekim_dict = {}
        cekim_tarihleri = {}
        if self.m_erken_cekim_liste:
            for kayit in self.m_erken_cekim_liste:
                kayit_tip = kayit[2] if len(kayit) == 3 else "yatirma"
                if kayit_tip == "cekme":
                    tarih_var, t_var = kayit[0], kayit[1]
                    tarih_str = tarih_var.get().strip()
                    t_str = t_var.get().strip()
                    if not tarih_str:
                        continue
                    if not t_str: t_str = "0"
                    try:
                        cekim_tarihi = datetime.datetime.strptime(tarih_str, "%d.%m.%Y").date()
                        t = float(t_str)
                        if t <= 0:
                            raise ValueError
                        delta_gun = (cekim_tarihi - tarih).days
                        if delta_gun <= 0:
                            if not sessiz:
                                self._hata_ve_temizle(None, "Çekim Tarihi Hatası",
                                                      "Para çekim tarihi başlangıç tarihinden sonra olmalıdır.", lbl)
                            return False
                        donem_no = math.ceil(delta_gun / vade)
                        if donem_no < 1: donem_no = 1
                        val = Decimal(str(-t))
                        if donem_no in erken_cekim_dict:
                            erken_cekim_dict[donem_no] += val
                        else:
                            erken_cekim_dict[donem_no] = val
                        if donem_no not in cekim_tarihleri or cekim_tarihi > cekim_tarihleri[donem_no]:
                            cekim_tarihleri[donem_no] = cekim_tarihi
                    except ValueError:
                        if not sessiz:
                            self._hata_ve_temizle(None, "Çekim Hatası",
                                                  "Para çekim satırlarını doğru doldurun (Tarih: GG.AA.YYYY, Tutar: pozitif sayı).",
                                                  lbl)
                        return False
                else:
                    d_var, t_var = kayit[0], kayit[1]
                    d_str = d_var.get().strip()
                    t_str = t_var.get().strip()
                    if not d_str: continue
                    if d_str and not t_str: t_str = "0"
                    try:
                        d = int(d_str)
                        t = float(t_str)
                        if d < 1: raise ValueError
                        if t <= 0: raise ValueError
                        val = Decimal(str(t))
                        if d in erken_cekim_dict:
                            erken_cekim_dict[d] += val
                        else:
                            erken_cekim_dict[d] = val
                    except:
                        if not sessiz:
                            self._hata_ve_temizle(None, "Hata", "Yatırma satırlarını doğru doldurun.", lbl)
                        return False

        birlesik_islemler = dict(ozel_islem_dict)
        for k, v in erken_cekim_dict.items():
            if k in birlesik_islemler:
                birlesik_islemler[k] += v
            else:
                birlesik_islemler[k] = v

        if not sessiz and birlesik_islemler:
            buyuk_donemler = [d for d in birlesik_islemler if d > N]
            if buyuk_donemler:
                self._hata_ve_temizle(None, "Özel İşlem Hatası",
                                      f"İşlem dönemi ({', '.join(map(str, sorted(buyuk_donemler)))}), belirlenen süreyi ({N}) aşıyor!",
                                      lbl)
                return False

        erken_cekim_var = any(len(k) == 3 and k[2] == "cekme" for k in self.m_erken_cekim_liste)
        if erken_cekim_var:
            erken_cekim_donemler = set()
            for kayit in self.m_erken_cekim_liste:
                if len(kayit) == 3 and kayit[2] == "cekme":
                    tarih_str = kayit[0].get().strip()
                    if tarih_str:
                        try:
                            cekim_tarihi = datetime.datetime.strptime(tarih_str, "%d.%m.%Y").date()
                            delta_gun = (cekim_tarihi - tarih).days
                            if delta_gun > 0:
                                donem_no = math.ceil(delta_gun / vade)
                                if donem_no < 1: donem_no = 1
                                erken_cekim_donemler.add(donem_no)
                        except Exception:
                            pass
        else:
            erken_cekim_donemler = set()
        tek_seferlik_mi = False
        sonuc = hesapla_mevduat(A0, faiz, P, N, stopaj, vade, tarih, erken_cekim_var=erken_cekim_var,
                                tek_seferlik=tek_seferlik_mi, ozel_islemler=birlesik_islemler, enflasyon=enf,
                                cekim_tarihleri=cekim_tarihleri, erken_cekim_donemler=erken_cekim_donemler)

        nom = sonuc["nominal"]
        reel = sonuc["reel"]
        p_min = sonuc["p_min"]
        bitis = sonuc["bitis_tarihi"].strftime("%d.%m.%Y")
        renk = r["ACCENT_GREEN"] if nom[-1] > 0 else r["ACCENT_RED"]

        t_brut = sonuc['toplam_brut_faiz']
        t_stop = sonuc['toplam_stopaj_kesinti']
        t_net = sonuc['toplam_net_faiz']
        oto_st = sonuc['oto_stopaj']
        net_son = Decimal(str(nom[-1]))
        brut_son_bakiye = net_son + t_stop

        veri_listesi = [
            ("Son Bakiye Tarihi:", bitis),
            ("Dizi Durumu:", sonuc['karakteristik']),
            ("Kritik Sınır P_min:", f"₺{p_min:,.2f} /Dönem"),
            ("", ""),
            ("Uygulanan Stopaj:", f"%{oto_st}"),
            ("Toplam Brüt Faiz:", f"₺{t_brut:,.2f}"),
            ("Toplam Stopaj Kes.:", f"₺{t_stop:,.2f}"),
            ("Toplam Net Faiz:", f"₺{t_net:,.2f}"),
            ("", ""),
            ("Brüt Son Bakiye:", f"₺{brut_son_bakiye:,.2f}"),
            ("Net Son Bakiye:", f"₺{net_son:,.2f}")
        ]
        if enf > 0:
            veri_listesi.append(("Reel Alım Gücü:", f"₺{reel[-1]:,.2f} (Enflasyon İskontolu)"))

        uyari_metinler = []

        if sonuc.get("bakiye_asimi_var"):
            detaylar = sonuc["bakiye_asimi_detay"]
            gruplu_uyarilar = []
            mevcut_grup = None

            for d in detaylar:
                if mevcut_grup is None:
                    mevcut_grup = {"baslangic": d["donem"], "bitis": d["donem"], "istenen": d["istenen"],
                                   "cekilen": d["cekilen"]}
                else:
                    if d["donem"] == mevcut_grup["bitis"] + 1 and d["istenen"] == mevcut_grup["istenen"] and d[
                        "cekilen"] == mevcut_grup["cekilen"]:
                        mevcut_grup["bitis"] = d["donem"]
                    else:
                        gruplu_uyarilar.append(mevcut_grup)
                        mevcut_grup = {"baslangic": d["donem"], "bitis": d["donem"], "istenen": d["istenen"],
                                       "cekilen": d["cekilen"]}

            if mevcut_grup:
                gruplu_uyarilar.append(mevcut_grup)

            bakiye_asimi_satirlari = []
            for g in gruplu_uyarilar:
                donem_str = f"{g['baslangic']}. Dönem" if g['baslangic'] == g[
                    'bitis'] else f"{g['baslangic']}-{g['bitis']}. Dönemler"
                bakiye_asimi_satirlari.append(
                    f"• {donem_str} -> İstenen: ₺{float(g['istenen']):,.2f} / Çekilebilen: ₺{float(g['cekilen']):,.2f}")

            if len(bakiye_asimi_satirlari) > 4:
                bakiye_asimi_satirlari = bakiye_asimi_satirlari[:3] + [
                    "• ... ve takip eden diğer dönemlerde de bakiye yetersiz kaldı."]

            uyari_metinler.append(
                "⚠️ BAKİYE YETERSİZ: Para çekimi mevcut bakiyeyi aştığı için içerdeki maksimum para çekildi:")
            uyari_metinler.extend(bakiye_asimi_satirlari)

        if sonuc["vade_bozuldu"]:
            vade_bozulanlar = sorted(list(sonuc["vade_bozulan_donemler"]))
            donemler_str = f"{vade_bozulanlar[0]}, {vade_bozulanlar[1]}, {vade_bozulanlar[2]}... dahil toplam {len(vade_bozulanlar)} farklı" if len(
                vade_bozulanlar) > 4 else ", ".join(map(str, vade_bozulanlar))
            uyari_metinler.append(
                f"❌ VADE BOZULDU: {donemler_str}. dönemde yapılan para çekiminden dolayı\n₺{sonuc['yanan_faiz']:,.2f} faiz YANDI! (Hesaplamalarda yanan faizler düşülmüştür)")
        elif "YAKINSAK" in sonuc["karakteristik"] and not sonuc.get("bakiye_asimi_var"):
            uyari_metinler.append(
                "⚠️ DİKKAT: Paranız zamanla eriyip tükenecek!\nÇekim miktarınız faiz getirinizi aşıyor.")

        if uyari_metinler:
            self._dinamik_yazdir(self.m_uyari_lbl, "\n".join(uyari_metinler), base_size=10, is_bold=True)
            self.m_uyari_lbl.config(fg=r["ACCENT_ORANGE"])
            if not sonuc.get("bakiye_asimi_var"):
                veri_listesi.append(("(Sıfır Hata Toleransıyla Hesaplanmıştır) ✓", ""))
        else:
            self.m_uyari_lbl.config(text="")
            veri_listesi.append(("(Sıfır Hata Toleransıyla Hesaplanmıştır) ✓", ""))

        self._sonuclari_yazdir(self.m_sonuc_grid_frame, veri_listesi, renk)

        self.tree_m.delete(*self.tree_m.get_children())

        gruplu_ekstre = []
        aktif_bos_grup = None

        for satir in sonuc["ekstre"]:
            if satir["faiz"] == Decimal('0.00') and satir["islem"] == Decimal('0.00') and satir["bakiye"] == Decimal(
                    '0.00') and not satir.get("kullanici_islemi", False):
                if aktif_bos_grup is None:
                    aktif_bos_grup = satir.copy()
                    aktif_bos_grup["orijinal_baslangic_donem"] = satir["donem"]
                    aktif_bos_grup["orijinal_baslangic_tarih"] = satir["baslangic_tarih"]
                else:
                    aktif_bos_grup["donem"] = f"{aktif_bos_grup['orijinal_baslangic_donem']}-{satir['donem']}"
                    aktif_bos_grup["tarih"] = satir["tarih"]
            else:
                if aktif_bos_grup is not None:
                    gruplu_ekstre.append(aktif_bos_grup)
                    aktif_bos_grup = None
                gruplu_ekstre.append(satir)

        if aktif_bos_grup is not None:
            gruplu_ekstre.append(aktif_bos_grup)

        for i, satir in enumerate(gruplu_ekstre):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            tgs = (tag, "vade_bozuldu") if satir.get("vade_bozuldu") else (tag,)
            baslangic_tarih_gosterim = satir.get("orijinal_baslangic_tarih", satir["baslangic_tarih"])
            self.tree_m.insert("", "end", values=(satir["donem"], baslangic_tarih_gosterim, satir["tarih"],
                                                  f"₺{satir['faiz']:,.2f}",
                                                  f"₺{satir['islem']:,.2f}", f"₺{satir['bakiye']:,.2f}",
                                                  f"₺{satir['reel_bakiye']:,.2f}"), tags=tgs)

        if erken_cekim_var and sonuc["vade_bozuldu"] and hasattr(self, 'm_tablo_not_lbl'):
            self.m_tablo_not_lbl.config(
                text="ℹ️  Turuncu renkle gösterilen satır, vadenin bozulduğu dönemdir. "
                     "Bir sonraki satır mevduat bozulduktan sonraki yeni dönemi gösterir.")
        elif hasattr(self, 'm_tablo_not_lbl'):
            self.m_tablo_not_lbl.config(text="")

        ciz_grafik(self.m_canvas, nom, reel if enf > 0 else [], renk, "Tasarruf Eğrisi", r, tree=self.tree_m,
                   vade_bozulan_donemler=sonuc.get("vade_bozulan_donemler", set()))
        self.m_hesaplandi = True

        if not sessiz and self.m_giris_acik:
            self._m_toggle_giris()

        if hasattr(self, 'btn_indir_csv_m'):
            self.btn_indir_csv_m.grid()
            self.btn_frame_m.grid_columnconfigure(0, weight=3)
            self.btn_frame_m.grid_columnconfigure(1, weight=1)
            self.btn_frame_m.grid_columnconfigure(2, weight=1)

        return True

    def _borc_hesapla(self, sessiz=False):
        r = self.renkler
        self.b_uyari_lbl.config(text="")
        lbl = self.b_uyari_lbl
        maks_kredi, maks_vade, mesaj, renk = self._bddk_sinirlari_hesapla()
        tip = self.b_kredi_tipi.get()

        if tip in ["Taşıt Kredisi", "Konut Kredisi"]:
            teminat_kontrol = self.b_teminat.get().replace('.', '').replace(',', '.')
            if teminat_kontrol in ["", ".", "-", "0"]:
                if not sessiz:
                    self._hata_ve_temizle(self.b_teminat, "Eksik Veri", f"Lütfen {tip} için geçerli bir değer giriniz.",
                                          lbl)
                return False

        faiz_str = self.b_faiz.get()
        if faiz_str in ["", ".", "-"]: faiz_str = "0"
        faiz = self._dogrula(faiz_str, 0, 100, "Kredi Faiz Oranı", lbl, sessiz)
        if faiz is None: return False

        vergi_str = self.b_vergi.get()
        if vergi_str in ["", ".", "-"]: vergi_str = "0"
        vergi = self._dogrula(vergi_str, 0, 100, "Vergi (KKDF)", lbl, sessiz)
        if vergi is None: return False

        tarih = self._tarih_dogrula(self.b_tarih.get(), lbl, sessiz)
        if tarih is None: return False

        baslangic_str = self.b_baslangic.get().replace('.', '').replace(',', '.')
        if baslangic_str in ["", ".", "-", "0"]:
            if not sessiz: self._hata_ve_temizle(self.b_baslangic, "Hata", "Talep Edilen Kredi 0 veya boş olamaz.", lbl)
            return False

        try:
            A0 = float(baslangic_str)
            if A0 <= 0:
                if not sessiz: self._hata_ve_temizle(self.b_baslangic, "Hata", "Talep Edilen Kredi 0 olamaz.", lbl)
                return False

            if maks_kredi != float('inf') and A0 > maks_kredi:
                A0 = float(maks_kredi)
                A0_dec = Decimal(str(maks_kredi)).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)
                A0_int = int(A0_dec)
                formatted_val = f"{A0_int:,}".replace(',', '.')
                decimal_part = str(A0_dec).split('.')[1]
                if decimal_part != "00": formatted_val += f",{decimal_part}"
                self.b_baslangic.set(formatted_val)

        except ValueError:
            if not sessiz: self._hata_ve_temizle(self.b_baslangic, "Hata", "'Toplam Borç' alanını sayısal girin.", lbl)
            return False

        odeme_str = self.b_odeme.get().replace('.', '').replace(',', '.')
        if odeme_str in ["", ".", "-"]: odeme_str = "0"
        try:
            P = float(odeme_str)
        except ValueError:
            if not sessiz: self._hata_ve_temizle(self.b_odeme, "Hata", "'Aylık Ödeme' formatı hatalı.", lbl)
            return False

        if not sessiz and P > A0 and A0 > 0:
            self._hata_ve_temizle(self.b_odeme, "Geçersiz Değer",
                                  f"Aylık ödeme (₺{P:,.2f}), toplam borcu (₺{A0:,.2f}) aşamaz.", lbl)
            return False

        hedef_vade = None
        if self.b_vade_var.get():
            vade_str = self.b_vade.get()
            if vade_str == "" or vade_str == "0":
                if not sessiz: self._hata_ve_temizle(self.b_vade, "Geçersiz Değer", "Vade Hedefi 1'den küçük olamaz.",
                                                     lbl)
                return False
            hedef_vade = self._dogrula_int(vade_str, 1, 120, "Vade Hedefi", lbl, sessiz)
            if hedef_vade is None: return False

            if not sessiz and hedef_vade > maks_vade:
                self._hata_ve_temizle(None, "BDDK Vade Limiti",
                                      f"Girilen vade ({hedef_vade} ay), BDDK yasal sınırını ({maks_vade} ay) aşıyor.",
                                      lbl)
                return False

        ara = {}
        for dv, tv in self.b_ara_odemeler_liste:
            d_val = dv.get().strip()
            t_val = tv.get().replace('.', '').replace(',', '.').strip()
            if d_val:
                if not t_val or t_val in [".", "-"]: t_val = "0"
                ara[int(d_val)] = Decimal(t_val)

        yeniden = self.b_yeniden_yapilandir_var.get()
        sonuc = hesapla_borc(A0, faiz, P, vergi, tarih, ara, hedef_vade=hedef_vade, yeniden_yapilandir=yeniden)
        kapanma_ay = sonuc["kapanma_ay"]

        if not self.b_vade_var.get() and not sonuc["iraksar"] and kapanma_ay > maks_vade:
            if not sessiz: self._hata_ve_temizle(None, "BDDK Vade Limiti",
                                                 f"Mevcut taksitle kredi {kapanma_ay} ayda bitiyor. Yasal sınır {maks_vade} aydır. Taksiti artırın.",
                                                 lbl)
            return False

        if sonuc["iraksar"]:
            renk = r["ACCENT_RED"]
            durum_aciklama = ("Yalnızca faiz ödeniyor,\nanapara eksilmiyor." if sonuc.get('iraksar_tip') == "sabit"
                              else "Ödeme faizi karşılamıyor,\nborç büyümeye devam eder.")

            veri_listesi = [
                ("Dizi Durumu:", sonuc['karakteristik']),
                ("Aylık Ödemeniz:", f"₺{P:,.2f}"),
                ("Aylık Faiz (Min):", f"₺{float(sonuc['p_min']):,.2f}"),
                ("Durum:", durum_aciklama),
                ("", ""),
                ("Toplam Ödenen:", f"₺{sonuc['toplam_odenen_nom']:,.2f}")
            ]

            uyari_metin = "⚠️ Aşağıda borcun 2 yılda (24 ay) nasıl büyüdüğünün logaritmik simülasyonu gösterilmiştir."
            self._dinamik_yazdir(self.b_uyari_lbl, uyari_metin, base_size=11, is_bold=True)
            self.b_uyari_lbl.config(fg=r["ACCENT_RED"])
        else:
            renk = r["ACCENT_GREEN"]
            bitis = sonuc["bitis_tarihi"].strftime("%d.%m.%Y")

            veri_listesi = [
                ("Borç Bitiş Tarihi:", bitis),
                ("Dizi Durumu:", sonuc['karakteristik']),
                ("", ""),
                ("Toplam Anapara:", f"₺{A0:,.2f}"),
                ("Toplam Ödenen:", f"₺{sonuc['toplam_odenen_nom']:,.2f}"),
                ("Toplam Faiz Yükü:", f"₺{float(sonuc['toplam_odenen_nom']) - A0:,.2f}")
            ]

            if sonuc.get("vade_kurtarmiyor"):
                uyari_metin = f"DİKKAT: Ödeme borcu hedeflenen sürede kapatmıyor!\n{hedef_vade}. ay sonunda ₺{sonuc['kalan_borc']:,.2f} borç kalıyor."
                self._dinamik_yazdir(self.b_uyari_lbl, uyari_metin, base_size=11, is_bold=True)
                self.b_uyari_lbl.config(fg=r["ACCENT_ORANGE"])
            elif hedef_vade is None:
                uyari_metin = f"✅ Borç {kapanma_ay} ayda ({kapanma_ay // 12} yıl {kapanma_ay % 12} ay) ödenerek kapatıldı."
                self._dinamik_yazdir(self.b_uyari_lbl, uyari_metin, base_size=11, is_bold=True)
                self.b_uyari_lbl.config(fg=r["ACCENT_GREEN"])
            else:
                if kapanma_ay < hedef_vade:
                    uyari_metin = f"✅ Yapılan ara ödemeler sayesinde borç, hedeflenen {hedef_vade} aydan daha ERKEN biterek {kapanma_ay}. ayda sıfırlanmıştır!"
                    self._dinamik_yazdir(self.b_uyari_lbl, uyari_metin, base_size=10, is_bold=True)
                    self.b_uyari_lbl.config(fg=r["ACCENT_GREEN"])
                elif kapanma_ay > hedef_vade:
                    uyari_metin = f"⚠️ Borç {kapanma_ay}. ayda sıfırlanmıştır."
                    self._dinamik_yazdir(self.b_uyari_lbl, uyari_metin, base_size=11, is_bold=True)
                    self.b_uyari_lbl.config(fg=r["ACCENT_ORANGE"])
                else:
                    uyari_metin = f"✅ Borç tam hedeflenen {hedef_vade}. ayda sıfırlanarak kapatıldı."
                    self._dinamik_yazdir(self.b_uyari_lbl, uyari_metin, base_size=11, is_bold=True)
                    self.b_uyari_lbl.config(fg=r["ACCENT_GREEN"])

        self._sonuclari_yazdir(self.b_sonuc_grid_frame, veri_listesi, renk)

        self.tree.delete(*self.tree.get_children())
        for i, satir in enumerate(sonuc["ekstre"]):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.tree.insert("", "end", values=(satir["taksit"], satir["tarih"], f"₺{satir['kalan_borc']:,.2f}",
                                                f"₺{satir['faiz']:,.2f}", f"₺{satir['anapara']:,.2f}",
                                                f"₺{satir.get('ara_odeme', Decimal('0.00')):,.2f}",
                                                f"₺{satir['tutar']:,.2f}"), tags=(tag,))

        grafik_baslik = "Borç Eğrisi (Iraksak Durum için Logaritmik)" if sonuc["iraksar"] else "Borç Eğrisi"
        ciz_grafik(self.b_canvas, sonuc.get("nominal", [A0]), [], renk, grafik_baslik, r, tree=self.tree,
                   log_olcek=sonuc["iraksar"])
        self.b_hesaplandi = True

        if not sessiz and self.b_giris_acik:
            self._b_toggle_giris()

        if hasattr(self, 'btn_indir_csv'):
            self.btn_indir_csv.grid()
            self.btn_frame_b.grid_columnconfigure(0, weight=3)
            self.btn_frame_b.grid_columnconfigure(1, weight=1)
            self.btn_frame_b.grid_columnconfigure(2, weight=1)

        return True


if __name__ == "__main__":
    kok = tk.Tk()
    uygulama = UygulamaGUI(kok)
    kok.withdraw()
    kok.update_idletasks()
    kok.deiconify()
    kok.lift()
    kok.attributes('-topmost', True)
    kok.after(200, lambda: kok.attributes('-topmost', False))
    kok.focus_force()

    try:
        kok.state('zoomed')
    except:
        try:
            kok.attributes('-zoomed', True)
        except:
            pass

    kok.mainloop()
